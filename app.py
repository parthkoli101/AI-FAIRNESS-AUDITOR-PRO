import os
import json
import math
import uuid
import time
import csv
import textwrap
import subprocess
import tempfile
import traceback
from io import StringIO
from datetime import datetime, timezone
from functools import wraps

from flask import (Flask, render_template, request, jsonify,
                   session, redirect, url_for, flash, make_response)
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
from werkzeug.exceptions import HTTPException, RequestEntityTooLarge
from dotenv import load_dotenv
import pandas as pd
import numpy as np
import requests as http_requests

try:
    from pymongo import MongoClient
    from bson import ObjectId
    MONGO_AVAILABLE = True
except ImportError:
    MONGO_AVAILABLE = False

try:
    import pdfplumber
    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False

try:
    import groq as groq_lib
    GROQ_LIB_AVAILABLE = True
except ImportError:
    GROQ_LIB_AVAILABLE = False

try:
    import google.generativeai as genai
    GEMINI_LIB_AVAILABLE = True
except ImportError:
    GEMINI_LIB_AVAILABLE = False

try:
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable
    from reportlab.platypus import KeepTogether
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    from reportlab.graphics.shapes import Drawing, Rect, String
    from reportlab.graphics import renderPDF
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

load_dotenv(override=True)

app = Flask(__name__)
app.secret_key = os.getenv("SECRET_KEY", "fallback_secret_key")
app.config["UPLOAD_FOLDER"] = "uploads"
app.config["MAX_CONTENT_LENGTH"] = 16 * 1024 * 1024

GROQ_API_KEY = os.getenv("GROQ_API_KEY", "").strip()
GEMINI_API_KEY = os.getenv("GEMINI_API_KEY", "").strip()
MONGO_URI = os.getenv("MONGO_URI", "mongodb://localhost:27017/")

LOCAL_DEV_ORIGINS = {
    "http://localhost:3000", "http://127.0.0.1:3000",
    "http://localhost:5000", "http://127.0.0.1:5000",
    "http://localhost:5500", "http://127.0.0.1:5500",
    "null",
}

# Configure Gemini if available
if GEMINI_LIB_AVAILABLE and GEMINI_API_KEY:
    genai.configure(api_key=GEMINI_API_KEY)

# ─── MongoDB Setup ──────────────────────────────────────────────────────────
db = None
if MONGO_AVAILABLE:
    try:
        client = MongoClient(MONGO_URI, serverSelectionTimeoutMS=3000)
        client.server_info()
        db = client["fairness_ai"]
        MONGO_AVAILABLE = True
    except Exception:
        MONGO_AVAILABLE = False

# ─── AI Provider Detection ──────────────────────────────────────────────────
def get_ai_provider():
    if GROQ_API_KEY and GROQ_LIB_AVAILABLE:
        return "groq"
    elif GEMINI_API_KEY and GEMINI_LIB_AVAILABLE:
        return "gemini"
    return None

def call_ai(prompt, max_tokens=1000, temperature=0.2):
    """Unified AI caller — tries Groq first, then Gemini."""
    provider = get_ai_provider()
    if not provider:
        return None
    try:
        if provider == "groq":
            client = groq_lib.Groq(api_key=GROQ_API_KEY)
            response = client.chat.completions.create(
                model="llama-3.3-70b-versatile",
                messages=[{"role": "user", "content": prompt}],
                max_tokens=max_tokens,
                temperature=temperature
            )
            return response.choices[0].message.content.strip()
        elif provider == "gemini":
            model = genai.GenerativeModel("gemini-2.0-flash")
            response = model.generate_content(prompt)
            return response.text.strip()
    except Exception:
        traceback.print_exc()
    return None

# ─── Auth Helpers ────────────────────────────────────────────────────────────
def get_groq_stress_error():
    if not GROQ_API_KEY:
        return "Groq is required for AI stress testing. Set GROQ_API_KEY locally and on Render."
    if not GROQ_LIB_AVAILABLE:
        return "Groq package is not installed. Ensure requirements.txt installs groq."
    return None


def call_groq_for_stress(prompt, max_tokens=700, temperature=0.2):
    """Stress testing is intentionally Groq-only."""
    err = get_groq_stress_error()
    if err:
        return None, err
    try:
        client = groq_lib.Groq(api_key=GROQ_API_KEY, timeout=18.0)
        response = client.chat.completions.create(
            model="llama-3.3-70b-versatile",
            messages=[{"role": "user", "content": prompt}],
            max_tokens=max_tokens,
            temperature=temperature
        )
        return response.choices[0].message.content.strip(), None
    except Exception as e:
        traceback.print_exc()
        return None, f"Groq stress analysis failed: {type(e).__name__}: {str(e)}"


def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if "user_id" not in session:
            if request.path.startswith("/api/"):
                return jsonify({"error": "Not authenticated. Please log in again."}), 401
            return redirect(url_for("login"))
        return f(*args, **kwargs)
    return decorated

# ─── Smart Text-to-DataFrame Parser ──────────────────────────────────────────
@app.after_request
def add_local_dev_cors_headers(response):
    origin = request.headers.get("Origin")
    if origin in LOCAL_DEV_ORIGINS and request.path.startswith("/api/"):
        response.headers["Access-Control-Allow-Origin"] = origin
        response.headers["Access-Control-Allow-Credentials"] = "true"
        response.headers["Access-Control-Allow-Headers"] = "Content-Type, Authorization"
        response.headers["Access-Control-Allow-Methods"] = "GET, POST, PUT, DELETE, OPTIONS"
    return response


@app.route("/api/<path:_path>", methods=["OPTIONS"])
def api_options(_path):
    return ("", 204)


def parse_text_as_dataset(text, source_hint="txt"):
    import re
    import io

    lines = [l.rstrip() for l in text.strip().splitlines() if l.strip()]
    if not lines:
        return None, "Empty file"

    for sep, name in [(",", "csv"), ("\t", "tsv"), ("|", "psv")]:
        candidate_lines = [l for l in lines if l.count(sep) >= 2]
        if len(candidate_lines) >= 5:
            try:
                joined = "\n".join(candidate_lines)
                df = pd.read_csv(io.StringIO(joined), sep=sep, engine="python", on_bad_lines="skip")
                if len(df) >= 3 and len(df.columns) >= 2:
                    return df, None
            except Exception:
                pass

    tag_pattern = re.compile(r"<([A-Za-z_][A-Za-z0-9_\-]*)(?:\s[^>]*)?>([^<]*)</\1>")
    matches = tag_pattern.findall(text)
    if len(matches) >= 10:
        tag_names = [m[0].lower() for m in matches]
        tag_counts = {}
        for t in tag_names:
            tag_counts[t] = tag_counts.get(t, 0) + 1
        all_matches = [(m[0].lower(), m[1].strip()) for m in tag_pattern.findall(text)]
        if all_matches:
            from collections import Counter
            field_counts = Counter(t for t, v in all_matches)
            n_records = max(field_counts.values())
            fields = [t for t, c in field_counts.items() if c >= max(1, n_records // 2)]
            if len(fields) >= 2:
                rows = []
                current = {}
                seen_fields = set()
                for tag, val in all_matches:
                    if tag in fields:
                        if tag in seen_fields:
                            rows.append({f: current.get(f, "") for f in fields})
                            current = {}
                            seen_fields = set()
                        current[tag] = val
                        seen_fields.add(tag)
                if current:
                    rows.append({f: current.get(f, "") for f in fields})
                if len(rows) >= 3:
                    df = pd.DataFrame(rows)
                    return df, None

    kv_pattern = re.compile(r"^([A-Za-z_][A-Za-z0-9_ \-/]*)[\s]*[:=]\s*(.+)$")
    blocks = []
    current_block = {}
    for line in lines:
        m = kv_pattern.match(line.strip())
        if m:
            key = m.group(1).strip().lower().replace(" ", "_")
            val = m.group(2).strip()
            current_block[key] = val
        else:
            if current_block:
                blocks.append(current_block)
                current_block = {}
    if current_block:
        blocks.append(current_block)

    if len(blocks) >= 3:
        from collections import Counter
        all_keys = Counter(k for b in blocks for k in b.keys())
        n = len(blocks)
        common_keys = [k for k, c in all_keys.items() if c >= max(2, n // 3)]
        if len(common_keys) >= 2:
            rows = [{k: b.get(k, "") for k in common_keys} for b in blocks]
            df = pd.DataFrame(rows)
            return df, None

    if len(lines) >= 4:
        header_re = re.compile(r"^[A-Za-z_][A-Za-z0-9_ ]*(\s{2,}[A-Za-z_][A-Za-z0-9_ ]*){2,}$")
        header_idx = None
        for i, line in enumerate(lines[:20]):
            if header_re.match(line.strip()):
                header_idx = i
                break
        if header_idx is not None:
            try:
                table_text = "\n".join(lines[header_idx:])
                df = pd.read_fwf(io.StringIO(table_text))
                if len(df) >= 3 and len(df.columns) >= 2:
                    return df, None
            except Exception:
                pass

    demographic_patterns = {
        "gender": re.compile(r"\b(male|female|man|woman|men|women|non.binary|nonbinary|transgender|gender)\b", re.I),
        "race": re.compile(r"\b(white|black|hispanic|latino|latina|asian|african|caucasian|indigenous|native|race|racial|ethnic)\b", re.I),
        "age": re.compile(r"\b(\d{2})\s*(?:year|yr)s?\s*old|\bage\s+(\d{2})|\b(young|old|senior|junior|elderly|youth|adult|teenager)\b", re.I),
        "disability": re.compile(r"\b(disabled|disability|handicap|impair|wheelchair|blind|deaf|autis|adhd)\b", re.I),
        "religion": re.compile(r"\b(christian|muslim|jewish|hindu|buddhist|religious|faith|church|mosque|temple)\b", re.I),
        "nationality": re.compile(r"\b(american|british|indian|chinese|mexican|canadian|nationality|immigrant|foreign|citizen)\b", re.I),
    }

    outcome_patterns = {
        "hired": re.compile(r"\b(hired|hire|accepted|approved|selected|admitted|passed|qualified)\b", re.I),
        "rejected": re.compile(r"\b(rejected|denied|refused|failed|disqualified|not selected|not hired)\b", re.I),
    }

    sentences = re.split(r"[.!?\n]+", text)
    rows = []
    for sent in sentences:
        sent = sent.strip()
        if len(sent) < 15:
            continue
        row = {}
        for attr, pat in demographic_patterns.items():
            matches_found = pat.findall(sent)
            if matches_found:
                val = matches_found[0]
                if isinstance(val, tuple):
                    val = next((v for v in val if v), attr)
                row[attr] = str(val).lower().strip()
        for outcome, pat in outcome_patterns.items():
            if pat.search(sent):
                row["outcome"] = outcome
                break
        if row and len(row) >= 2:
            rows.append(row)

    if len(rows) >= 5:
        df = pd.DataFrame(rows)
        df = df.fillna("unknown")
        return df, None

    term_counts = {}
    for attr, pat in demographic_patterns.items():
        found = pat.findall(text)
        if found:
            for match in found:
                if isinstance(match, tuple):
                    match = next((v for v in match if v), attr)
                key = f"{attr}_{str(match).lower().strip()}"
                term_counts[key] = term_counts.get(key, 0) + 1

    if term_counts:
        rows = [{"term": k, "frequency": v, "category": k.split("_")[0]}
                for k, v in term_counts.items()]
        df = pd.DataFrame(rows)
        return df, None

    return None, "Could not extract structured data from this file"


# ─── File Loader ─────────────────────────────────────────────────────────────
def load_uploaded_file(file):
    filename = secure_filename(file.filename)
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""

    if ext == "csv":
        try:
            df = pd.read_csv(file)
            return "dataframe", df
        except Exception as e:
            return "error", str(e)

    elif ext == "json":
        try:
            raw = json.load(file)
            if isinstance(raw, list):
                df = pd.json_normalize(raw)
            elif isinstance(raw, dict):
                df = pd.json_normalize(raw)
            else:
                return "error", "Unsupported JSON structure"
            return "dataframe", df
        except Exception as e:
            return "error", str(e)

    elif ext in ("xlsx", "xls"):
        try:
            df = pd.read_excel(file)
            return "dataframe", df
        except Exception as e:
            return "error", str(e)

    elif ext == "xml":
        try:
            import xml.etree.ElementTree as ET
            raw_bytes = file.read()
            text = raw_bytes.decode("utf-8", errors="replace")
            try:
                root = ET.fromstring(raw_bytes)
                from collections import Counter
                child_tags = Counter(child.tag for child in root)
                if child_tags:
                    record_tag = child_tags.most_common(1)[0][0]
                    rows = []
                    for record in root.findall(record_tag):
                        row = {}
                        for field in record:
                            row[field.tag.lower()] = field.text or ""
                        for attr_name, attr_val in record.attrib.items():
                            row[attr_name.lower()] = attr_val
                        if row:
                            rows.append(row)
                    if len(rows) >= 2:
                        return "dataframe", pd.DataFrame(rows)
                rows = []
                for child in root:
                    row = {"tag": child.tag.lower()}
                    row["value"] = child.text or ""
                    for sub in child:
                        row[sub.tag.lower()] = sub.text or ""
                    for attr_name, attr_val in child.attrib.items():
                        row[attr_name.lower()] = attr_val
                    rows.append(row)
                if rows:
                    return "dataframe", pd.DataFrame(rows)
            except ET.ParseError:
                pass
            df, err = parse_text_as_dataset(text, "xml")
            if df is not None and len(df) >= 3:
                return "dataframe", df
            return "text", text
        except Exception as e:
            return "error", str(e)

    elif ext in ("html", "htm"):
        try:
            raw_bytes = file.read()
            text = raw_bytes.decode("utf-8", errors="replace")
            import io
            try:
                tables = pd.read_html(io.StringIO(text))
                if tables:
                    best = max(tables, key=lambda t: len(t))
                    if len(best) >= 2 and len(best.columns) >= 2:
                        return "dataframe", best
            except Exception:
                pass
            import re
            clean = re.sub(r"<script[^>]*>.*?</script>", " ", text, flags=re.DOTALL | re.I)
            clean = re.sub(r"<style[^>]*>.*?</style>", " ", clean, flags=re.DOTALL | re.I)
            clean = re.sub(r"<[^>]+>", " ", clean)
            clean = re.sub(r"\s+", " ", clean).strip()
            df, err = parse_text_as_dataset(clean, "html")
            if df is not None and len(df) >= 3:
                return "dataframe", df
            return "text", clean
        except Exception as e:
            return "error", str(e)

    elif ext == "txt":
        try:
            text = file.read().decode("utf-8", errors="replace")
            df, err = parse_text_as_dataset(text, "txt")
            if df is not None and len(df) >= 3:
                return "dataframe", df
            return "text", text
        except Exception as e:
            return "error", str(e)

    elif ext == "pdf":
        try:
            if PDF_AVAILABLE:
                with pdfplumber.open(file) as pdf:
                    pages = [page.extract_text() or "" for page in pdf.pages]
                text = "\n".join(pages)
            else:
                text = file.read().decode("utf-8", errors="replace")
            df, err = parse_text_as_dataset(text, "pdf")
            if df is not None and len(df) >= 3:
                return "dataframe", df
            return "text", text
        except Exception as e:
            return "error", str(e)

    else:
        try:
            text = file.read().decode("utf-8", errors="replace")
            df, err = parse_text_as_dataset(text, ext)
            if df is not None and len(df) >= 3:
                return "dataframe", df
            return "text", text
        except Exception:
            return "error", f"Unsupported file type: {ext}"

# ─── MongoDB Normalizer ───────────────────────────────────────────────────────
def normalize_for_mongo(obj):
    if isinstance(obj, dict):
        return {str(k): normalize_for_mongo(v) for k, v in obj.items()}
    elif isinstance(obj, list):
        return [normalize_for_mongo(i) for i in obj]
    elif isinstance(obj, (np.integer,)):
        return int(obj)
    elif isinstance(obj, (np.floating,)):
        if math.isnan(obj) or math.isinf(obj):
            return None
        return float(obj)
    elif isinstance(obj, np.ndarray):
        return obj.tolist()
    elif isinstance(obj, float):
        if math.isnan(obj) or math.isinf(obj):
            return None
        return obj
    return obj

# ─── Protected Attribute Detection ───────────────────────────────────────────
PROTECTED_KEYWORDS = [
    "gender", "sex", "race", "ethnicity", "disability", "religion",
    "nationality", "age", "marital", "pregnancy", "origin"
]

def detect_protected_attributes(df):
    found = []
    for col in df.columns:
        col_lower = str(col).lower()
        for kw in PROTECTED_KEYWORDS:
            if kw in col_lower:
                found.append(col)
                break
    return found

def bin_age_column(df):
    df = df.copy()
    for col in df.columns:
        if "age" in str(col).lower() and pd.api.types.is_numeric_dtype(df[col]):
            bins = [0, 25, 35, 45, 60, 200]
            labels = ["18-25", "26-35", "36-45", "46-60", "60+"]
            df[col] = pd.cut(df[col], bins=bins, labels=labels, right=True)
    return df

TARGET_KEYWORDS = [
    "hired", "approved", "selected", "accepted", "rejected",
    "loan_approved", "admitted", "decision", "outcome",
    "prediction", "result", "label"
]

POSITIVE_VALUES = {
    "1", "yes", "true", "approved", "accept", "accepted",
    "selected", "hired", "positive", "pass"
}

def detect_target_column(df):
    for col in df.columns:
        if str(col).lower().strip() in TARGET_KEYWORDS:
            return col
    return None

def convert_to_binary(series):
    if pd.api.types.is_numeric_dtype(series):
        return pd.to_numeric(series, errors="coerce").fillna(0).apply(lambda x: 1 if x > 0 else 0)
    return series.astype(str).str.lower().str.strip().apply(
        lambda x: 1 if x in POSITIVE_VALUES else 0
    )


def _resolve_column_name(df, name):
    if not name:
        return None
    if name in df.columns:
        return name
    lookup = {str(c).lower().strip(): c for c in df.columns}
    return lookup.get(str(name).lower().strip())


def _coerce_dataframe_types(df):
    df = df.copy()
    for col in df.columns:
        if str(col).startswith("_"):
            continue
        series = df[col]
        if pd.api.types.is_numeric_dtype(series):
            continue
        as_num = pd.to_numeric(series.astype(str).str.replace(",", "", regex=False), errors="coerce")
        non_null = int(series.notna().sum())
        if non_null > 0 and int(as_num.notna().sum()) / non_null >= 0.8:
            df[col] = as_num
    return df


def _normalize_prediction_outcome(value):
    if value is None:
        return 0
    if isinstance(value, bool):
        return 1 if value else 0
    if isinstance(value, (int, float)):
        return 1 if float(value) > 0 else 0
    text = str(value).lower().strip()
    if text in POSITIVE_VALUES:
        return 1
    if text in {"0", "no", "false", "rejected", "reject", "deny", "denied", "negative", "fail", "failed"}:
        return 0
    try:
        return 1 if float(text) > 0 else 0
    except Exception:
        return 0


def _align_profile_row(profile, feature_columns, template_df):
    row = {}
    for col in feature_columns:
        if col not in profile:
            row[col] = np.nan
            continue
        val = profile[col]
        if val == "" or val is None:
            row[col] = np.nan
            continue
        if col in template_df.columns and pd.api.types.is_numeric_dtype(template_df[col]):
            try:
                row[col] = float(val)
            except (TypeError, ValueError):
                row[col] = pd.to_numeric(val, errors="coerce")
        else:
            row[col] = str(val)
    return pd.DataFrame([row])


def _profile_attr_value(profile, protected_attr):
    if protected_attr in profile:
        return str(profile.get(protected_attr))
    lookup = {str(k).lower().strip(): v for k, v in profile.items()}
    return str(lookup.get(str(protected_attr).lower().strip(), "unknown"))


def _group_approval_rates(results, protected_attr):
    counts = {}
    for item in results:
        profile = item.get("profile", {})
        attr_val = _profile_attr_value(profile, protected_attr)
        if attr_val not in counts:
            counts[attr_val] = {"approved": 0, "total": 0}
        counts[attr_val]["total"] += 1
        outcome = item.get("result")
        if "error" not in item and _normalize_prediction_outcome(outcome) == 1:
            counts[attr_val]["approved"] += 1
    rates = {}
    for attr_val, data in counts.items():
        total = data["total"]
        rates[attr_val] = round(data["approved"] / total, 4) if total > 0 else 0.0
    return rates


def _rates_to_bias_score(rates):
    if len(rates) < 2:
        return 0.0
    values = list(rates.values())
    if max(values) <= 0:
        return 0.0
    spd = max(values) - min(values)
    di = min(values) / max(values)
    return compute_bias_score({
        "statistical_parity_difference": {"counterfactual": spd},
        "disparate_impact": {"counterfactual": di},
    })

def compute_bias_score(metrics):
    spd_dict = metrics.get("statistical_parity_difference", {})
    di_dict = metrics.get("disparate_impact", {})

    if isinstance(spd_dict, (int, float)):
        spd_dict = {"_": spd_dict}
    if isinstance(di_dict, (int, float)):
        di_dict = {"_": di_dict}

    spd_scores, di_scores = [], []

    for spd in spd_dict.values():
        if spd is not None:
            spd_scores.append(min(max(float(spd), 0.0), 1.0))
    for di in di_dict.values():
        if di is not None:
            di_val = min(max(float(di), 0.0), 1.0)
            di_bias = 1.0 - di_val
            di_scores.append(di_bias)

    worst_spd = max(spd_scores) if spd_scores else 0.0
    worst_di = max(di_scores) if di_scores else 0.0
    combined = max(worst_spd, worst_di)

    if combined == 0:
        base_score = 0.0
    elif combined <= 0.05:
        base_score = combined * 140
    elif combined <= 0.10:
        base_score = 7 + (combined - 0.05) * 160
    elif combined <= 0.20:
        base_score = 15 + (combined - 0.10) * 130
    elif combined <= 0.35:
        base_score = 28 + (combined - 0.20) * 113
    elif combined <= 0.50:
        base_score = 45 + (combined - 0.35) * 100
    elif combined <= 0.70:
        base_score = 60 + (combined - 0.50) * 90
    elif combined <= 0.90:
        base_score = 78 + (combined - 0.70) * 70
    else:
        base_score = 92 + (combined - 0.90) * 70

    missing_pct = metrics.get("missing_pct", {})
    missing_penalty = min(sum(1 for v in missing_pct.values() if v > 20) * 2, 5)
    alert_penalty = min(metrics.get("alert_count", 0) * 2, 5)
    final_score = base_score + missing_penalty + alert_penalty

    complete_exclusion = False
    for group_rates in list(metrics.get("outcome_bias", {}).values()) + list(metrics.get("group_rates", {}).values()):
        if isinstance(group_rates, dict):
            for v in group_rates.values():
                if v == 0.0:
                    complete_exclusion = True
                    break

    if not complete_exclusion:
        final_score = min(final_score, 97.0)

    return round(min(final_score, 100.0), 1)


def score_to_risk(score):
    if score < 15:
        return "LOW"
    elif score < 35:
        return "MEDIUM"
    elif score < 60:
        return "HIGH"
    return "CRITICAL"


def compute_audit_confidence(metrics):
    score = 100
    if not metrics.get("protected_attributes"):
        score -= 30
    if not metrics.get("distributions") and not metrics.get("group_rates"):
        score -= 25
    if not metrics.get("total_rows") or metrics.get("total_rows", 0) < 100:
        score -= 20
    if score >= 80:
        return "HIGH"
    elif score >= 55:
        return "MEDIUM"
    return "LOW"


def compute_text_metrics(text, filename="document"):
    import re
    from collections import Counter

    DEMO_PATTERNS = {
        "gender": {
            "male":       re.compile(r"\b(male|man|men|boy|his|he)\b", re.I),
            "female":     re.compile(r"\b(female|woman|women|girl|her|she)\b", re.I),
            "non-binary": re.compile(r"\b(non.binary|nonbinary|they|them|enby|genderqueer)\b", re.I),
        },
        "race": {
            "white":     re.compile(r"\b(white|caucasian|european)\b", re.I),
            "black":     re.compile(r"\b(black|african.american|afro)\b", re.I),
            "hispanic":  re.compile(r"\b(hispanic|latino|latina|latinx)\b", re.I),
            "asian":     re.compile(r"\b(asian|chinese|japanese|korean|indian)\b", re.I),
            "indigenous":re.compile(r"\b(indigenous|native|aboriginal|first.nation)\b", re.I),
        },
        "age_group": {
            "18-25":  re.compile(r"\b(1[89]|2[0-5])\s*(?:year|yr)s?\s*old|\byouth\b|\bteenager\b|\byoung adult\b", re.I),
            "26-35":  re.compile(r"\b(2[6-9]|3[0-5])\s*(?:year|yr)s?\s*old|\byoung professional\b", re.I),
            "36-45":  re.compile(r"\b(3[6-9]|4[0-5])\s*(?:year|yr)s?\s*old|\bmid.career\b", re.I),
            "46-60":  re.compile(r"\b(4[6-9]|5[0-9]|60)\s*(?:year|yr)s?\s*old|\bsenior employee\b", re.I),
            "60+":    re.compile(r"\b(6[1-9]|[789]\d)\s*(?:year|yr)s?\s*old|\belderly\b|\bretir", re.I),
        },
        "disability": {
            "disabled":      re.compile(r"\b(disab|handicap|impair|wheelchair|blind|deaf|autis|adhd|dyslexia)\w*", re.I),
            "non-disabled":  re.compile(r"\b(able.bodied|neurotypical|no disability|without disab)\b", re.I),
        },
        "religion": {
            "christian": re.compile(r"\b(christian|church|catholic|protestant|baptist|evangelical)\b", re.I),
            "muslim":    re.compile(r"\b(muslim|islam|mosque|hijab|halal)\b", re.I),
            "jewish":    re.compile(r"\b(jewish|jew|synagogue|kosher|rabbi)\b", re.I),
            "hindu":     re.compile(r"\b(hindu|temple|diwali|brahmin)\b", re.I),
            "other":     re.compile(r"\b(buddhist|sikh|atheist|agnostic|secular)\b", re.I),
        },
        "nationality": {
            "domestic":      re.compile(r"\b(citizen|national|domestic|born here|local)\b", re.I),
            "international": re.compile(r"\b(immigrant|foreign|overseas|visa|migrant|expat)\b", re.I),
        },
    }

    OUTCOME_PATTERNS = {
        "positive": re.compile(
            r"\b(hired|accepted|approved|selected|admitted|passed|qualified|promoted|"
            r"successful|granted|offered|recruited|shortlisted|advanced)\b", re.I),
        "negative": re.compile(
            r"\b(rejected|denied|refused|failed|disqualified|dismissed|terminated|"
            r"not selected|not hired|not approved|turned down|laid off)\b", re.I),
    }

    group_counts = {}
    for attr, groups in DEMO_PATTERNS.items():
        counts = {}
        for group_name, pat in groups.items():
            c = len(pat.findall(text))
            if c > 0:
                counts[group_name] = c
        if counts:
            group_counts[attr] = counts

    total_mentions = sum(sum(g.values()) for g in group_counts.values())
    total_chars = len(text)
    total_words = len(text.split())

    sentences = re.split(r"[.!?\n]+", text)
    outcome_by_group = {}

    for sent in sentences:
        sent = sent.strip()
        if len(sent) < 10:
            continue
        pos = bool(OUTCOME_PATTERNS["positive"].search(sent))
        neg = bool(OUTCOME_PATTERNS["negative"].search(sent))
        if not pos and not neg:
            continue
        for attr, groups in DEMO_PATTERNS.items():
            for group_name, pat in groups.items():
                if pat.search(sent):
                    if attr not in outcome_by_group:
                        outcome_by_group[attr] = {}
                    if group_name not in outcome_by_group[attr]:
                        outcome_by_group[attr][group_name] = {"pos": 0, "neg": 0, "total": 0}
                    if pos:
                        outcome_by_group[attr][group_name]["pos"] += 1
                    if neg:
                        outcome_by_group[attr][group_name]["neg"] += 1
                    outcome_by_group[attr][group_name]["total"] += 1

    metrics = {}
    metrics["total_rows"] = total_words
    metrics["total_columns"] = len(group_counts)
    metrics["columns"] = list(group_counts.keys())
    metrics["source"] = "text_extraction"
    metrics["file_type"] = filename.rsplit(".", 1)[-1].lower() if "." in filename else "txt"
    metrics["text_stats"] = {
        "total_characters": total_chars,
        "total_words": total_words,
        "total_sentences": len([s for s in sentences if len(s.strip()) > 10]),
        "demographic_mentions": total_mentions,
    }

    metrics["protected_attributes"] = list(group_counts.keys())

    distributions = {}
    for attr, counts in group_counts.items():
        total = sum(counts.values())
        if total > 0:
            distributions[attr] = {g: round(c / total, 4) for g, c in counts.items()}
    metrics["distributions"] = distributions
    metrics["missing_values"] = {}
    metrics["missing_pct"] = {}

    outcome_bias = {}
    spd_values = {}
    di_values = {}
    alerts = []
    target_col = "document_outcome"

    for attr, groups in outcome_by_group.items():
        group_rates = {}
        for group_name, counts in groups.items():
            total = counts["total"]
            if total > 0:
                rate = round(counts["pos"] / total, 4)
                group_rates[group_name] = rate

        if len(group_rates) >= 2:
            rates = list(group_rates.values())
            max_rate = max(rates)
            min_rate = min(rates)
            spd = round(max_rate - min_rate, 4)
            di = round(min_rate / max_rate, 4) if max_rate > 0 else 0
            outcome_bias[attr] = group_rates
            spd_values[attr] = spd
            di_values[attr] = di
            if spd > 0.10 or di < 0.80:
                alerts.append({
                    "type": "outcome_bias",
                    "column": attr,
                    "target": target_col,
                    "statistical_parity_difference": spd,
                    "disparate_impact": di,
                    "detail": f"Outcome bias in '{attr}': SPD={spd}, DI={di}",
                    "severity": "HIGH" if (spd > 0.20 or di < 0.60) else "MEDIUM"
                })

    for attr, dist in distributions.items():
        if dist:
            max_val = max(dist.values())
            if max_val > 0.70:
                dominant = max(dist, key=dist.get)
                alerts.append({
                    "type": "imbalance",
                    "column": attr,
                    "dominant_group": dominant,
                    "pct": round(max_val * 100, 1),
                    "severity": "HIGH" if max_val > 0.85 else "MEDIUM"
                })

    metrics["outcome_bias"] = outcome_bias
    metrics["statistical_parity_difference"] = spd_values
    metrics["disparate_impact"] = di_values
    metrics["target_column"] = target_col if outcome_bias else None
    metrics["alerts"] = alerts
    metrics["alert_count"] = len(alerts)
    metrics["bias_score"] = compute_bias_score(metrics)
    metrics["risk_level"] = score_to_risk(metrics["bias_score"])
    metrics["audit_confidence"] = "MEDIUM" if total_mentions >= 20 else "LOW"

    return metrics


def generate_text_graph_data(metrics):
    graphs = []
    for attr, dist in metrics.get("distributions", {}).items():
        if dist:
            graphs.append({
                "type": "bar",
                "title": f"{attr.replace('_', ' ').title()} Mentions",
                "labels": list(dist.keys()),
                "data": [round(v * 100, 1) for v in dist.values()],
                "attr": attr
            })
    for attr, group_rates in metrics.get("outcome_bias", {}).items():
        if group_rates:
            graphs.append({
                "type": "bar",
                "title": f"Positive Outcome Rate by {attr.replace('_', ' ').title()}",
                "labels": list(group_rates.keys()),
                "data": [round(v * 100, 1) for v in group_rates.values()],
                "attr": attr
            })
    ts = metrics.get("text_stats", {})
    if ts:
        graphs.append({
            "type": "bar",
            "title": "Document Overview",
            "labels": ["Words", "Sentences", "Demo Mentions"],
            "data": [
                ts.get("total_words", 0),
                ts.get("total_sentences", 0),
                ts.get("demographic_mentions", 0)
            ],
            "attr": "text_stats"
        })
    return graphs


def generate_alert_explanation_groq(alert, target_col=None):
    attr = alert.get("column", "unknown")
    alert_type = alert.get("type", "")
    severity = alert.get("severity", "HIGH")

    if alert_type == "outcome_bias":
        spd = alert.get("statistical_parity_difference", 0)
        di = alert.get("disparate_impact", 0)
        target = target_col or alert.get("target", "the outcome")
        spd_pct = round(float(spd) * 100, 1)
        di_pct = round(float(di) * 100, 1)
        rule80_status = "FAILS the 80% rule" if float(di) < 0.80 else "passes the 80% rule"

        prompt = f"""You are an AI fairness expert writing a bias alert explanation for a hiring/selection AI system dashboard.

DETECTED BIAS FACTS (use these exact numbers, do not change them):
- Protected attribute: {attr}
- Affected outcome column: {target}
- Statistical Parity Difference (SPD): {spd} — one group is {spd_pct}% more likely to receive a positive outcome
- Disparate Impact ratio (DI): {di} — the disadvantaged group is only {di_pct}% as likely to be selected as the dominant group
- 80% Rule status: {rule80_status}
- Severity: {severity}

Write a bias explanation with EXACTLY these 3 parts, in plain text, no markdown:

WHERE BIAS EXISTS:
[One clear sentence: state exactly which attribute ('{attr}') shows bias affecting which outcome ('{target}'). Name the disparity direction.]

WHY BIAS EXISTS:
[Two sentences: explain what the SPD and DI values mean in plain terms for a hiring context. Use the exact numbers. Explain what real-world impact this has on the disadvantaged group.]

WHAT TO DO:
[One sentence: one specific, actionable recommendation to reduce this bias.]

Rules: Use only the facts above. No invented numbers. No legal references. No markdown. Plain text only."""

    elif alert_type == "imbalance":
        pct = alert.get("pct", 0)
        dominant = alert.get("dominant_group", "unknown")

        prompt = f"""You are an AI fairness expert writing a dataset imbalance alert for a hiring AI dashboard.

DETECTED IMBALANCE FACTS (use these exact numbers):
- Protected attribute: {attr}
- Dominant group: {dominant} — makes up {pct}% of the dataset
- All other groups share only {round(100 - float(pct), 1)}% of the dataset combined
- Severity: {severity}

Write an imbalance explanation with EXACTLY these 3 parts, in plain text, no markdown:

WHERE BIAS EXISTS:
[One sentence: state that '{attr}' is severely imbalanced — '{dominant}' dominates at {pct}%.]

WHY BIAS EXISTS:
[Two sentences: explain how this imbalance causes a trained AI model to develop biased behavior. State that underrepresented groups will be systematically disadvantaged in model decisions.]

WHAT TO DO:
[One sentence: specific recommendation — oversample minority groups, collect more data, or apply class weights during training.]

Rules: Use only the facts above. No invented numbers. Plain text only. No markdown."""
    else:
        return _fallback_alert_explanation(alert)

    result = call_ai(prompt, max_tokens=300, temperature=0.2)
    if result:
        return result
    return _fallback_alert_explanation(alert)


def _fallback_alert_explanation(alert):
    attr = alert.get("column", "unknown")
    alert_type = alert.get("type", "")
    if alert_type == "outcome_bias":
        spd = alert.get("statistical_parity_difference", 0)
        di = alert.get("disparate_impact", 0)
        target = alert.get("target", "the outcome")
        spd_pct = round(float(spd) * 100, 1)
        di_pct = round(float(di) * 100, 1)
        return (
            f"WHERE BIAS EXISTS:\n"
            f"Outcome bias detected in '{target}' for the protected attribute '{attr}'.\n\n"
            f"WHY BIAS EXISTS:\n"
            f"One group is {spd_pct}% more likely to receive a positive outcome (SPD={spd}). "
            f"The disadvantaged group is only {di_pct}% as likely to be selected as the dominant group (DI={di}), "
            f"which {'fails' if float(di) < 0.80 else 'passes'} the 80% fairness rule.\n\n"
            f"WHAT TO DO:\n"
            f"Apply fairness constraints during model training and audit decision thresholds per group."
        )
    elif alert_type == "imbalance":
        pct = alert.get("pct", 0)
        dominant = alert.get("dominant_group", "unknown")
        return (
            f"WHERE BIAS EXISTS:\n"
            f"Severe representation imbalance in '{attr}': '{dominant}' makes up {pct}% of the dataset.\n\n"
            f"WHY BIAS EXISTS:\n"
            f"With {pct}% of data belonging to '{dominant}', the AI model will learn patterns biased toward this group. "
            f"Underrepresented groups account for only {round(100 - float(pct), 1)}% of training data, causing systematic disadvantage.\n\n"
            f"WHAT TO DO:\n"
            f"Oversample underrepresented groups or apply class weighting to balance the training distribution."
        )
    return (
        f"WHERE BIAS EXISTS:\nBias alert detected in column '{attr}'.\n\n"
        f"WHY BIAS EXISTS:\nThis attribute shows potential discriminatory patterns requiring review.\n\n"
        f"WHAT TO DO:\nConduct a detailed audit of this attribute's distribution and impact on outcomes."
    )


def compute_graph_analysis(graph):
    labels = graph.get("labels", [])
    values = graph.get("data", graph.get("values", []))
    title = graph.get("title", "")

    if not labels or not values or len(labels) != len(values):
        return None

    safe_vals, safe_labels = [], []
    for l, v in zip(labels, values):
        try:
            safe_vals.append(float(v))
            safe_labels.append(str(l))
        except (TypeError, ValueError):
            pass

    if len(safe_vals) < 2:
        return None

    max_val = max(safe_vals)
    min_val = min(safe_vals)
    max_idx = safe_vals.index(max_val)
    min_idx = safe_vals.index(min_val)
    total = sum(safe_vals)
    avg_val = total / len(safe_vals) if safe_vals else 0

    if max_val > 0:
        relative_disparity = (max_val - min_val) / max_val
    else:
        relative_disparity = 0.0

    absolute_gap = max_val - min_val
    is_distribution = total > 50
    is_rate = total <= 2.0

    if is_rate:
        risk_flag = "HIGH" if absolute_gap > 0.20 else ("MEDIUM" if absolute_gap > 0.10 else "LOW")
    elif is_distribution:
        risk_flag = "HIGH" if relative_disparity > 0.50 else ("MEDIUM" if relative_disparity > 0.30 else "LOW")
    else:
        risk_flag = "HIGH" if absolute_gap > 30 else ("MEDIUM" if absolute_gap > 15 else "LOW")

    return {
        "title": title,
        "dominant_group": safe_labels[max_idx],
        "dominant_value": round(max_val, 4),
        "lowest_group": safe_labels[min_idx],
        "lowest_value": round(min_val, 4),
        "absolute_gap": round(absolute_gap, 4),
        "relative_disparity_pct": round(relative_disparity * 100, 1),
        "average_value": round(avg_val, 4),
        "risk_flag": risk_flag,
        "is_balanced": relative_disparity < 0.15,
        "all_groups": {l: round(v, 4) for l, v in zip(safe_labels, safe_vals)},
        "group_count": len(safe_labels),
    }


def explain_graph_with_groq(graph_summary):
    analysis = compute_graph_analysis(graph_summary)
    if analysis is None:
        return "Insufficient data for analysis."

    dom = analysis["dominant_group"]
    dom_val = analysis["dominant_value"]
    low = analysis["lowest_group"]
    low_val = analysis["lowest_value"]
    gap = analysis["absolute_gap"]
    rel_pct = analysis["relative_disparity_pct"]
    risk = analysis["risk_flag"]
    title = analysis["title"]
    is_balanced = analysis["is_balanced"]
    groups_str = ", ".join([f"{g}: {v}" for g, v in analysis["all_groups"].items()])

    if is_balanced and risk == "LOW":
        risk_conclusion = f"BALANCED — no significant bias detected. The gap of {rel_pct}% is within the acceptable range."
        affected_line = "No significantly affected group — distribution is within acceptable range."
        risk_line = "Low risk — no significant fairness concern detected in this chart."
        rec_line = "Continue monitoring group distributions over time to ensure balance is maintained."
    elif risk == "MEDIUM":
        risk_conclusion = f"MODERATE disparity detected. {dom} is overrepresented. Gap of {rel_pct}% warrants monitoring."
        affected_line = f"{low} is the most affected group with a lower rate/representation."
        risk_line = f"Medium risk — {dom} has a {rel_pct}% advantage over {low}. This may affect fairness."
        rec_line = f"Monitor group rates closely and consider rebalancing {low} representation in the dataset."
    else:
        risk_conclusion = f"HIGH disparity detected. {dom} significantly dominates {low} with a {rel_pct}% gap."
        affected_line = f"{low} is the most affected group with value {low_val} vs {dom}'s {dom_val}."
        risk_line = f"High risk — {rel_pct}% disparity may lead to systematic bias against {low} in hiring/selection."
        rec_line = f"Rebalance the dataset: increase {low} representation or apply fairness constraints during retraining."

    prompt = f"""You are an AI fairness auditor writing a brief chart explanation.

PYTHON-COMPUTED FACTS — USE THESE EXACT VALUES, DO NOT CHANGE THEM:
Chart: {title}
Groups: {groups_str}
Highest group: {dom} = {dom_val}
Lowest group: {low} = {low_val}
Gap: {gap} ({rel_pct}% relative difference)
Python risk assessment: {risk}
Python conclusion: {risk_conclusion}

YOUR TASK: Write natural-language explanations using ONLY the facts above.

STRICT RULES (violation = failure):
1. Use ONLY the numbers shown above. NEVER invent or change any number.
2. If Python says LOW/BALANCED -> you MUST NOT flag it as risky. Say no significant bias.
3. If Python says HIGH -> explain disparity using the exact values above.
4. No civic, political, public-participation, or social language.
5. No elections, voting, outreach, or community programs.
6. Hiring/selection fairness context ONLY.
7. 1-2 sentences per section. Do NOT repeat points.

Return EXACTLY this format, no markdown, no extra text:

Key Finding:
[One sentence stating what the chart shows using exact values.]

Evidence:
[One sentence citing exact group values from the facts above.]

Affected Group:
[{affected_line}]

Risk:
[{risk_line}]

Recommendation:
[{rec_line}]"""

    result = call_ai(prompt, max_tokens=400, temperature=0.2)
    if result:
        return result

    return f"""Key Finding:
{"The " + title + " shows a balanced distribution with no significant bias detected." if is_balanced and risk == "LOW" else f"The {title} shows a {risk.lower()} disparity — {dom} ({dom_val}) significantly exceeds {low} ({low_val})."}

Evidence:
{"Groups: " + groups_str + f". Relative gap is only {rel_pct}%, within the acceptable range." if is_balanced else f"Exact values: {groups_str}. The absolute gap is {gap} ({rel_pct}% relative difference)."}

Affected Group:
{affected_line}

Risk:
{risk_line}

Recommendation:
{rec_line}"""


# ─── Pre-Training Metrics ─────────────────────────────────────────────────────
def compute_pretraining_metrics(df):
    metrics = {}
    metrics["total_rows"] = int(len(df))
    metrics["total_columns"] = int(len(df.columns))
    metrics["columns"] = list(df.columns)

    df_binned = bin_age_column(df)
    protected_attrs = detect_protected_attributes(df_binned)
    metrics["protected_attributes"] = protected_attrs

    missing = df.isnull().sum()
    missing_pct = (missing / len(df) * 100).round(2)
    metrics["missing_values"] = {col: int(missing[col]) for col in df.columns if missing[col] > 0}
    metrics["missing_pct"] = {col: float(missing_pct[col]) for col in df.columns if missing_pct[col] > 0}

    distributions = {}
    alerts = []
    for attr in protected_attrs:
        if attr not in df_binned.columns:
            continue
        col_data = df_binned[attr].dropna()
        if pd.api.types.is_numeric_dtype(col_data) and "age" not in attr.lower():
            continue
        vc = col_data.value_counts(normalize=True).round(4)
        dist = {str(k): float(v) for k, v in vc.items()}
        distributions[attr] = dist
        if vc.max() > 0.85:
            alerts.append({
                "type": "imbalance",
                "column": attr,
                "dominant_group": str(vc.idxmax()),
                "pct": float(round(vc.max() * 100, 1))
            })

    metrics["distributions"] = distributions

    target_col = detect_target_column(df)
    outcome_bias = {}
    spd_values = {}
    di_values = {}

    if target_col:
        df_outcome = df_binned.copy()
        df_outcome["_target_binary"] = convert_to_binary(df[target_col])
        metrics["target_column"] = target_col

        for attr in protected_attrs:
            if attr not in df_outcome.columns:
                continue
            col_data = df_outcome[attr].dropna()
            if pd.api.types.is_numeric_dtype(col_data) and "age" not in attr.lower():
                continue
            try:
                grp = df_outcome.groupby(attr)["_target_binary"].mean().round(4)
                group_rates = {str(k): float(v) for k, v in grp.items()}
                if len(group_rates) >= 2:
                    rates = list(group_rates.values())
                    max_rate = max(rates)
                    min_rate = min(rates)
                    spd = round(max_rate - min_rate, 4)
                    di = round(min_rate / max_rate, 4) if max_rate > 0 else 0
                    outcome_bias[attr] = group_rates
                    spd_values[attr] = spd
                    di_values[attr] = di
                    if spd > 0.10 or di < 0.80:
                        alerts.append({
                            "type": "outcome_bias",
                            "column": attr,
                            "target": target_col,
                            "statistical_parity_difference": spd,
                            "disparate_impact": di,
                            "detail": f"Outcome bias detected in '{target_col}' for '{attr}': SPD={spd}, DI={di}",
                            "severity": "HIGH" if (spd > 0.20 or di < 0.60) else "MEDIUM"
                        })
            except Exception:
                pass

    metrics["outcome_bias"] = outcome_bias
    metrics["statistical_parity_difference"] = spd_values
    metrics["disparate_impact"] = di_values
    metrics["alerts"] = alerts
    metrics["alert_count"] = len(alerts)
    metrics["bias_score"] = compute_bias_score(metrics)
    metrics["risk_level"] = score_to_risk(metrics["bias_score"])
    metrics["audit_confidence"] = compute_audit_confidence(metrics)
    return metrics


# ─── Post-Training Metrics ────────────────────────────────────────────────────
def compute_posttraining_metrics(df, prediction_col, label_col=None):
    metrics = {}
    metrics["total_predictions"] = int(len(df))
    metrics["prediction_col"] = prediction_col

    if prediction_col not in df.columns:
        metrics["error"] = f"Prediction column '{prediction_col}' not found"
        return metrics

    preds = df[prediction_col]
    unique_vals = preds.dropna().unique()

    binary_map = {}
    if set(str(v).lower() for v in unique_vals).issubset({"yes", "no", "true", "false", "1", "0", "approved", "rejected", "accept", "deny", "positive", "negative"}):
        positive_vals = {"yes", "true", "1", "approved", "accept", "positive"}
        binary_map = {v: 1 if str(v).lower() in positive_vals else 0 for v in unique_vals}
        df = df.copy()
        df["_pred_binary"] = df[prediction_col].map(lambda x: binary_map.get(x, 1 if str(x).lower() in positive_vals else 0))
    else:
        df = df.copy()
        try:
            df["_pred_binary"] = pd.to_numeric(df[prediction_col], errors="coerce").fillna(0).astype(int)
        except Exception:
            df["_pred_binary"] = 1

    pos_rate = float(df["_pred_binary"].mean())
    metrics["positive_rate"] = round(pos_rate, 4)
    outcome_counts = df["_pred_binary"].value_counts().to_dict()
    metrics["outcome_counts"] = {str(k): int(v) for k, v in outcome_counts.items()}

    df_binned = bin_age_column(df)
    protected_attrs = detect_protected_attributes(df_binned)
    metrics["protected_attributes"] = protected_attrs

    group_rates = {}
    spd_values = {}
    di_values = {}

    for attr in protected_attrs:
        if attr not in df_binned.columns:
            continue
        col_data = df_binned[attr].dropna()
        if pd.api.types.is_numeric_dtype(col_data) and "age" not in attr.lower():
            continue
        try:
            grp = df_binned.groupby(attr)["_pred_binary"].mean().round(4)
            gr = {str(k): float(v) for k, v in grp.items()}
            group_rates[attr] = gr
            if len(gr) >= 2:
                rates = list(gr.values())
                spd = max(rates) - min(rates)
                spd_values[attr] = round(spd, 4)
                di = min(rates) / max(rates) if max(rates) > 0 else None
                di_values[attr] = round(di, 4) if di is not None else None
        except Exception:
            pass

    metrics["group_rates"] = group_rates
    metrics["statistical_parity_difference"] = spd_values
    metrics["disparate_impact"] = di_values

    rule_80 = {}
    for attr, di in di_values.items():
        if di is not None:
            rule_80[attr] = "PASS" if di >= 0.8 else "FAIL"
    metrics["rule_80"] = rule_80

    if label_col and label_col in df.columns:
        try:
            from sklearn.metrics import accuracy_score, precision_score, recall_score, f1_score
            labels = df[label_col]
            label_map = {v: 1 if str(v).lower() in {"yes", "true", "1", "approved", "accept", "positive"} else 0 for v in labels.dropna().unique()}
            y_true = labels.map(lambda x: label_map.get(x, int(x) if str(x).isdigit() else 0))
            y_pred = df["_pred_binary"]
            valid = ~(y_true.isna() | y_pred.isna())
            y_true = y_true[valid].astype(int)
            y_pred = y_pred[valid].astype(int)
            metrics["accuracy"] = round(float(accuracy_score(y_true, y_pred)), 4)
            metrics["precision"] = round(float(precision_score(y_true, y_pred, zero_division=0)), 4)
            metrics["recall"] = round(float(recall_score(y_true, y_pred, zero_division=0)), 4)
            metrics["f1"] = round(float(f1_score(y_true, y_pred, zero_division=0)), 4)
        except Exception:
            pass

    metrics["bias_score"] = compute_bias_score(metrics)
    metrics["risk_level"] = score_to_risk(metrics["bias_score"])
    metrics["audit_confidence"] = compute_audit_confidence(metrics)
    return metrics


def generate_graph_data(metrics, mode):
    graphs = []
    if mode == "pre":
        dists = metrics.get("distributions", {})
        for attr, dist in dists.items():
            graphs.append({
                "type": "bar",
                "title": f"{attr} Distribution",
                "labels": list(dist.keys()),
                "data": [round(v * 100, 1) for v in dist.values()],
                "attr": attr
            })
        missing = metrics.get("missing_pct", {})
        if missing:
            graphs.append({
                "type": "bar",
                "title": "Missing Values (%)",
                "labels": list(missing.keys()),
                "data": list(missing.values()),
                "attr": "missing"
            })
    elif mode == "post":
        group_rates = metrics.get("group_rates", {})
        for attr, gr in group_rates.items():
            graphs.append({
                "type": "bar",
                "title": f"Approval Rate by {attr}",
                "labels": list(gr.keys()),
                "data": [round(v * 100, 1) for v in gr.values()],
                "attr": attr
            })
        outcome = metrics.get("outcome_counts", {})
        if outcome:
            graphs.append({
                "type": "pie",
                "title": "Outcome Distribution",
                "labels": ["Positive", "Negative"],
                "data": [outcome.get("1", 0), outcome.get("0", 0)],
                "attr": "outcome"
            })
        spd = metrics.get("statistical_parity_difference", {})
        if spd:
            graphs.append({
                "type": "bar",
                "title": "Statistical Parity Difference by Attribute",
                "labels": list(spd.keys()),
                "data": [v if v is not None else 0 for v in spd.values()],
                "attr": "spd"
            })
    return graphs


# ─── Sandbox helpers ──────────────────────────────────────────────────────────
def explain_sandbox_with_groq(approval_by_group, bias_score, protected_attr):
    if not approval_by_group:
        return ""
    rates_str = ", ".join([f"{g}: {int(v*100)}%" for g, v in approval_by_group.items()])
    max_rate = max(approval_by_group.values())
    min_rate = min(approval_by_group.values())
    max_group = max(approval_by_group, key=approval_by_group.get)
    min_group = min(approval_by_group, key=approval_by_group.get)
    gap = max_rate - min_rate
    is_balanced = gap < 0.10
    risk_level = score_to_risk(bias_score)

    prompt = f"""You are an AI fairness auditor reviewing a counterfactual bias test.

PYTHON-COMPUTED FACTS (use exactly as given):
- Protected attribute: {protected_attr}
- Group approval rates: {rates_str}
- Highest group: {max_group} at {int(max_rate*100)}%
- Lowest group: {min_group} at {int(min_rate*100)}%
- Absolute gap: {round(gap*100, 1)}%
- Bias score: {bias_score}/100
- Risk: {risk_level}
- Balance: {"BALANCED — gap under 10%, no significant counterfactual bias" if is_balanced else f"IMBALANCED — {round(gap*100,1)}% gap detected"}

RULES:
1. Use only these numbers. Never invent values.
2. If balanced -> MUST say no significant bias.
3. 1-2 sentences per section only.

Return this format exactly:

Key Finding:
[State whether changing {protected_attr} changed predictions.]

Evidence:
[Exact approval rates for each group.]

Affected Group:
[{min_group} if gap >= 10%, else "No affected group — gap is within acceptable range."]

Risk:
[State risk with reason.]

Recommendation:
[One specific next step.]"""

    return call_ai(prompt, max_tokens=400, temperature=0.2) or ""


def generate_sandbox_with_groq(protected_attr, domain="hiring"):
    prompt = f"""Generate a biased synthetic Python predict(row) function for testing a fairness auditor in {domain}.
Protected attribute: {protected_attr}
Rules: Only def predict(row):. No imports. Return 1=approved, 0=rejected. Different outcomes for GroupA/B/C/D. Max 12 lines.
Respond ONLY in JSON with no markdown, no backticks, no explanation:
{{"code": "def predict(row):\\n    ...", "sample_json": {{"{protected_attr}": "GroupA", "income": 55000, "score": 70}}, "protected_attr": "{protected_attr}"}}"""

    result = call_ai(prompt, max_tokens=400, temperature=0.3)
    if result:
        try:
            start = result.find("{")
            end = result.rfind("}") + 1
            if start >= 0 and end > start:
                return json.loads(result[start:end])
        except Exception:
            pass

    return {
        "code": f"""def predict(row):
    group = row.get('{protected_attr}', 'GroupA')
    income = float(row.get('income', 50000))
    rates = {{'GroupA': 0.8, 'GroupB': 0.6, 'GroupC': 0.4, 'GroupD': 0.2}}
    threshold = rates.get(group, 0.5)
    return 1 if income > 50000 * (1 / threshold) else 0""",
        "sample_json": {protected_attr: "GroupA", "income": 55000, "score": 70},
        "protected_attr": protected_attr
    }


def explain_optimizer_with_groq(optimizer_result):
    before = optimizer_result.get("before_score", 0)
    after = optimizer_result.get("after_score", 0)
    improvement = optimizer_result.get("improvement", 0)
    attr = optimizer_result.get("protected_attr", "attribute")
    before_dist = optimizer_result.get("before_dist", {})
    after_dist = optimizer_result.get("after_dist", {})
    before_str = ", ".join([f"{g}: {round(v*100,1)}%" for g, v in before_dist.items()])
    after_str = ", ".join([f"{g}: {round(v*100,1)}%" for g, v in after_dist.items()])

    prompt = f"""You are an AI fairness auditor reviewing a bias optimization simulation.

PYTHON FACTS (use exactly):
- Attribute: {attr}
- Before distribution: {before_str}
- After distribution: {after_str}
- Bias score before: {before}/100 -> after: {after}/100
- Improvement: {improvement} points

RULES: Use only these numbers. No civic/political language. 1-2 sentences each.

Return exactly:

Bias Found:
[Explain original imbalance with exact values.]

Optimization Effect:
[Explain score change.]

Affected Group:
[Groups remaining underrepresented after optimization.]

Reliability Check:
[Is simulation realistic or unstable?]

Recommendation:
[One practical next step.]"""

    return call_ai(prompt, max_tokens=400, temperature=0.2) or ""


# ─── AI Report Generator ──────────────────────────────────────────────────────
def generate_ai_report(metrics, mode):
    provider = get_ai_provider()
    if not provider:
        return {"error": "No AI provider configured. Set GROQ_API_KEY or GEMINI_API_KEY in .env"}

    mode_label = {"pre": "Pre-Training Dataset Audit", "post": "Post-Training Model Audit", "appeal": "Appeal Engine Report"}.get(mode, "Fairness Audit")

    prompt = f"""You are a senior AI fairness auditor for a hiring/selection AI system.
Audit type: {mode_label}
COMPUTED METRICS:
{json.dumps(metrics, indent=2)}

STRICT RULES:
- Use ONLY the provided metrics.
- Do NOT invent numbers, causes, laws, or external facts.
- Do NOT use civic outreach, public participation, or political language.
- Do NOT blame protected groups.
- Focus on dataset quality, approval rates, SPD, DI, 80% rule, and bias risk.
- If SPD > 0.10 -> meaningful disparity. If SPD > 0.20 -> severe disparity.
- If DI < 0.80 -> fails 80% fairness rule.
- If bias_score >= 60 -> final verdict must be CRITICAL.
- If key metrics are missing -> INCONCLUSIVE.

Return EXACTLY these 9 sections:

1. EXECUTIVE SUMMARY
2. DATASET/MODEL COMPOSITION
3. MOST AFFECTED GROUPS
4. BIAS RISK ANALYSIS
5. FAIRNESS METRIC FINDINGS
6. LEGAL & ETHICAL IMPLICATIONS
7. FUTURE RISK ASSESSMENT
8. RECOMMENDATIONS
- rebalance underrepresented groups
- review labels for historical bias
- tune decision thresholds per fairness policy
- retrain using fairness constraints
- monitor group-level approval rates
- require human review for high-impact decisions

9. FINAL VERDICT
Return only one: PASS, INCONCLUSIVE, FAIL, CRITICAL.
Then 2-3 sentences explaining why.

Each section: 3-5 sentences max. Plain text only. No markdown headings."""

    result = call_ai(prompt, max_tokens=2000, temperature=0.3)
    if result:
        return {"report": result, "provider": get_ai_provider()}
    return {"error": "AI report generation failed"}


# ─── What-If Simulations ──────────────────────────────────────────────────────
def run_what_if_pretraining(df, protected_attr, desired_balance):
    if protected_attr not in df.columns:
        return {"error": f"Column '{protected_attr}' not found"}
    df = bin_age_column(df)
    col = df[protected_attr].dropna()
    before_dist = col.value_counts(normalize=True).round(4).to_dict()
    before_dist = {str(k): float(v) for k, v in before_dist.items()}
    temp_metrics_before = compute_pretraining_metrics(df)
    before_score = temp_metrics_before["bias_score"]

    groups = list(before_dist.keys())
    n = len(groups)
    if n == 0:
        return {"error": "No groups found"}
    target = float(desired_balance) / 100
    leftover = 1.0 - target
    after_dist = {}
    majority = max(before_dist, key=before_dist.get)
    for g in groups:
        if g == majority:
            after_dist[g] = round(target, 4)
        else:
            after_dist[g] = round(leftover / (n - 1), 4) if n > 1 else 0.0

    # Resample df to match after_dist while simulating outcome-aware rebalancing
    df_sim_list = []
    total_rows = len(df)
    target_col = detect_target_column(df)
    
    # Calculate majority positive selection rate
    majority_pos_rate = 0.5
    df_target_bin = df.copy()
    if target_col:
        df_target_bin["_target_binary"] = convert_to_binary(df[target_col])
        maj_df = df_target_bin[df_target_bin[protected_attr].astype(str) == str(majority)]
        if len(maj_df) > 0:
            majority_pos_rate = float(maj_df["_target_binary"].mean())
    else:
        # Fallback target column identification if name-matching didn't catch it
        for c in df.columns:
            if c != protected_attr:
                df_target_bin["_target_binary"] = convert_to_binary(df[c])
                target_col = c
                maj_df = df_target_bin[df_target_bin[protected_attr].astype(str) == str(majority)]
                if len(maj_df) > 0:
                    majority_pos_rate = float(maj_df["_target_binary"].mean())
                break

    for g in groups:
        g_df = df_target_bin[df_target_bin[protected_attr].astype(str) == str(g)]
        if len(g_df) == 0:
            continue
        
        # Desired count of rows for this group in the simulation
        g_target_count = max(5, int(after_dist[g] * total_rows))
        
        if "_target_binary" in df_target_bin.columns:
            # Adjust selection rate towards majority based on balance factor
            g_pos_df = g_df[g_df["_target_binary"] == 1]
            g_neg_df = g_df[g_df["_target_binary"] == 0]
            current_rate = len(g_pos_df) / len(g_df) if len(g_df) > 0 else 0.0
            
            balance_factor = 1.0
            if n > 1:
                ideal_share = 1.0 / n
                current_share = after_dist[majority]
                if current_share > ideal_share:
                    balance_factor = max(0.0, 1.0 - (current_share - ideal_share) / (1.0 - ideal_share))
            
            # The more balanced the representation, the more selection rate difference is mitigated
            target_rate = current_rate + (majority_pos_rate - current_rate) * balance_factor
            
            target_pos_count = max(0, min(g_target_count, int(target_rate * g_target_count)))
            target_neg_count = max(0, g_target_count - target_pos_count)
            
            resampled_parts = []
            if target_pos_count > 0 and len(g_pos_df) > 0:
                resampled_parts.append(g_pos_df.sample(n=target_pos_count, replace=True, random_state=42))
            elif target_pos_count > 0 and len(g_df) > 0:
                resampled_parts.append(g_df.sample(n=target_pos_count, replace=True, random_state=42))
                
            if target_neg_count > 0 and len(g_neg_df) > 0:
                resampled_parts.append(g_neg_df.sample(n=target_neg_count, replace=True, random_state=42))
            elif target_neg_count > 0 and len(g_df) > 0:
                resampled_parts.append(g_df.sample(n=target_neg_count, replace=True, random_state=42))
                
            if resampled_parts:
                df_sim_list.append(pd.concat(resampled_parts, ignore_index=True))
        else:
            df_sim_list.append(g_df.sample(n=g_target_count, replace=True, random_state=42))

    if df_sim_list:
        df_sim = pd.concat(df_sim_list, ignore_index=True)
        if "_target_binary" in df_sim.columns:
            # Revert temp column name back to original target column values
            if target_col and target_col in df_sim.columns:
                df_sim[target_col] = df_sim["_target_binary"]
            df_sim = df_sim.drop(columns=["_target_binary"])
    else:
        df_sim = df.copy()

    temp_metrics_after = compute_pretraining_metrics(df_sim)
    after_score = temp_metrics_after["bias_score"]
    improvement = max(0, round(before_score - after_score, 1))
    return {
        "protected_attr": protected_attr,
        "before_dist": before_dist,
        "after_dist": after_dist,
        "before_score": before_score,
        "after_score": after_score,
        "improvement": improvement,
        "groups": groups
    }


def _resolve_prediction_column(df, prediction_col=None):
    if prediction_col and prediction_col in df.columns:
        return prediction_col
    detected = detect_target_column(df)
    if detected and detected in df.columns:
        return detected
    return df.columns[-1] if len(df.columns) else None


def _apply_pred_binary(df, prediction_col):
    df = df.copy()
    if not prediction_col or prediction_col not in df.columns:
        return df, None

    preds = df[prediction_col]
    unique_vals = preds.dropna().unique()
    positive_vals = {"yes", "true", "1", "approved", "accept", "positive"}
    categorical_vals = {
        "yes", "no", "true", "false", "1", "0",
        "approved", "rejected", "accept", "deny", "positive", "negative",
    }
    if set(str(v).lower() for v in unique_vals).issubset(categorical_vals):
        binary_map = {v: 1 if str(v).lower() in positive_vals else 0 for v in unique_vals}
        df["_pred_binary"] = df[prediction_col].map(
            lambda x: binary_map.get(x, 1 if str(x).lower() in positive_vals else 0)
        )
    else:
        try:
            df["_pred_binary"] = pd.to_numeric(df[prediction_col], errors="coerce").fillna(0).astype(int)
        except Exception:
            df["_pred_binary"] = 1
    return df, prediction_col


def run_what_if_posttraining(df, threshold, fairness_weight, protected_attr, prediction_col=None):
    prediction_col = _resolve_prediction_column(df, prediction_col)
    if not prediction_col:
        return {"error": "No prediction column found for optimization"}

    df, prediction_col = _apply_pred_binary(df, prediction_col)
    if "_pred_binary" not in df.columns:
        return {"error": f"Prediction column '{prediction_col}' could not be parsed"}

    df = bin_age_column(df)
    before_dist = {}
    if protected_attr and protected_attr in df.columns:
        grp_before = df.groupby(df[protected_attr].astype(str))["_pred_binary"].mean().round(4)
        before_dist = {str(k): float(v) for k, v in grp_before.items()}

    before_rate = float(df["_pred_binary"].mean())
    df2 = df.copy()

    # 1. Simulate threshold shift: adjust overall approval rates
    total_len = len(df2)
    current_pos = int(df2["_pred_binary"].sum())
    target_rate = before_rate
    if threshold > 0.5:
        target_rate *= (1.0 - (threshold - 0.5) * 2.0)
    elif threshold < 0.5:
        target_rate += (1.0 - target_rate) * ((0.5 - threshold) * 2.0)
    target_pos = int(target_rate * total_len)

    if target_pos < current_pos:
        pos_indices = df2[df2["_pred_binary"] == 1].index
        flips = current_pos - target_pos
        if len(pos_indices) >= flips and flips > 0:
            df2.loc[df2.index.isin(pos_indices.to_series().sample(n=flips, random_state=42)), "_pred_binary"] = 0
    elif target_pos > current_pos:
        neg_indices = df2[df2["_pred_binary"] == 0].index
        flips = target_pos - current_pos
        if len(neg_indices) >= flips and flips > 0:
            df2.loc[df2.index.isin(neg_indices.to_series().sample(n=flips, random_state=42)), "_pred_binary"] = 1

    # 2. Simulate fairness weight shift: raise disadvantaged groups toward the max group rate
    if protected_attr and protected_attr in df2.columns:
        grp_rates = df2.groupby(df2[protected_attr].astype(str))["_pred_binary"].mean().to_dict()
        if len(grp_rates) >= 2:
            max_rate = max(grp_rates.values())
            for g, r in grp_rates.items():
                if r < max_rate:
                    g_mask = df2[protected_attr].astype(str) == str(g)
                    g_indices = df2[g_mask].index
                    g_neg_indices = df2[g_mask & (df2["_pred_binary"] == 0)].index
                    target_g_rate = r + (max_rate - r) * fairness_weight
                    target_g_pos = int(target_g_rate * len(g_indices))
                    current_g_pos = int(df2.loc[g_indices, "_pred_binary"].sum())
                    needed_flips = max(0, target_g_pos - current_g_pos)
                    if needed_flips > 0 and len(g_neg_indices) >= needed_flips:
                        df2.loc[
                            df2.index.isin(g_neg_indices.to_series().sample(n=needed_flips, random_state=42)),
                            "_pred_binary",
                        ] = 1

    after_rate = float(df2["_pred_binary"].mean())
    after_dist = {}
    if protected_attr and protected_attr in df2.columns:
        grp_after = df2.groupby(df2[protected_attr].astype(str))["_pred_binary"].mean().round(4)
        after_dist = {str(k): float(v) for k, v in grp_after.items()}

    before_metrics = compute_posttraining_metrics(df, "_pred_binary")
    after_metrics = compute_posttraining_metrics(df2, "_pred_binary")

    return {
        "protected_attr": protected_attr,
        "before_dist": before_dist,
        "after_dist": after_dist,
        "before_rates": before_dist,
        "after_rates": after_dist,
        "before_score": before_metrics.get("bias_score", 0),
        "after_score": after_metrics.get("bias_score", 0),
        "before_positive_rate": round(before_rate, 4),
        "after_positive_rate": round(after_rate, 4),
        "improvement": max(0, round(before_metrics.get("bias_score", 0) - after_metrics.get("bias_score", 0), 1)),
        "threshold": threshold,
        "fairness_weight": fairness_weight,
        "groups": list(before_dist.keys()),
        "prediction_col": prediction_col,
    }


# ─── Stress Testing ───────────────────────────────────────────────────────────
def run_blackbox_stress_test_api(api_url, token, sample_json, protected_attr):
    if not api_url:
        return {"error": "API URL required"}
    try:
        base_payload = json.loads(sample_json) if isinstance(sample_json, str) else sample_json
    except Exception:
        return {"error": "Invalid sample JSON"}

    headers = {"Content-Type": "application/json"}
    if token:
        headers["Authorization"] = f"Bearer {token}"

    synthetic_groups = ["GroupA", "GroupB", "GroupC", "GroupD"]
    results = []
    for group in synthetic_groups:
        payload = dict(base_payload)
        payload[protected_attr] = group
        try:
            resp = http_requests.post(api_url, json=payload, headers=headers, timeout=10)
            decision = resp.json()
        except Exception as e:
            decision = {"error": str(e)}
        results.append({"group": group, "payload": payload, "response": decision})

    approval_by_group = {}
    for r in results:
        resp = r["response"]
        approved = 0
        if isinstance(resp, dict):
            for key in ["approved", "decision", "result", "prediction", "label"]:
                if key in resp:
                    approved = _normalize_prediction_outcome(resp[key])
                    break
        approval_by_group[r["group"]] = approved

    rates = list(approval_by_group.values())
    if len(rates) >= 2 and max(rates) > 0:
        spd = max(rates) - min(rates)
        di = min(rates) / max(rates)
        temp_metrics = {"statistical_parity_difference": {"g": spd}, "disparate_impact": {"g": di}}
        bias_score = compute_bias_score(temp_metrics)
    else:
        bias_score = _rates_to_bias_score(approval_by_group) if len(approval_by_group) >= 2 else 0

    return {
        "results": results,
        "approval_by_group": approval_by_group,
        "bias_score": round(bias_score, 1),
        "risk_level": score_to_risk(bias_score),
        "counterfactual_rate": round(bias_score, 1)
    }


def run_blackbox_stress_test_sandbox(code, sample_json, protected_attr):
    if not code or not code.strip():
        return {"error": "No code provided"}
    try:
        base_payload = json.loads(sample_json) if isinstance(sample_json, str) else sample_json
    except Exception:
        return {"error": "Invalid sample JSON"}

    synthetic_profiles = [
        _sanitize_profile({**base_payload, protected_attr: g})
        for g in ["GroupA", "GroupB", "GroupC", "GroupD", "GroupE", "GroupF", "", None]
    ]

    result_data = run_sandbox_on_profiles(code, synthetic_profiles)
    if result_data.get("error"):
        return result_data

    results = result_data.get("results", [])
    approval_by_group = {}
    for r in results:
        profile = r.get("profile", {})
        group = _profile_attr_value(profile, protected_attr)
        if group == "unknown":
            group = str(profile.get(protected_attr, "unknown"))
        approval_by_group[group] = _normalize_prediction_outcome(r.get("result", 0)) if "error" not in r else 0

    bias_score = _rates_to_bias_score(approval_by_group) if len(approval_by_group) >= 2 else 0
    groq_explanation = explain_sandbox_with_groq(approval_by_group, bias_score, protected_attr)

    return {
        "results": results,
        "approval_by_group": approval_by_group,
        "bias_score": round(bias_score, 1),
        "risk_level": score_to_risk(bias_score),
        "counterfactual_rate": round(bias_score, 1),
        "profiles_tested": len(synthetic_profiles),
        "groq_explanation": groq_explanation
    }


# ─── Appeal Engine ────────────────────────────────────────────────────────────
def run_appeal_engine(document_text, policy_text, domain):
    provider = get_ai_provider()
    if not provider:
        return {"error": "No AI provider configured"}

    prompt = f"""You are an AI appeal analyst for a hiring/selection decision.
Domain: {domain}
USER DOCUMENT:
{document_text[:3000]}
POLICY / CRITERIA:
{policy_text[:3000]}
STRICT RULES:
- Use only the user document and policy text.
- Do NOT invent achievements, skills, reasons, or missing requirements.
- Be neutral and non-discriminatory.
- Focus on policy match, evidence gaps, and appeal quality.

Return EXACTLY this format:
FIT SCORE: [0-100]
MATCHED REQUIREMENTS:
- [requirement + evidence]
MISSING REQUIREMENTS:
- [requirement not found]
LIKELY REJECTION REASONS:
- [policy-based reasons only]
POSSIBLE FAIRNESS CONCERNS:
- [only if relevant; otherwise "none identified"]
IMPROVEMENT PLAN:
- [specific improvements]
APPEAL RECOMMENDATION:
[Strong Appeal / Moderate Appeal / Weak Appeal / No Appeal Recommended]
APPEAL SUMMARY:
[2-3 sentences.]
No markdown. No invented facts."""

    result = call_ai(prompt, max_tokens=2000, temperature=0.3)
    if not result:
        return {"error": "AI generation failed"}

    fit_score = 50
    for line in result.split("\n"):
        if line.strip().startswith("FIT SCORE:"):
            try:
                fit_score = int(line.split(":")[1].strip().split()[0])
            except Exception:
                pass

    return {"report": result, "fit_score": fit_score, "provider": get_ai_provider()}


# ─── PDF Export ───────────────────────────────────────────────────────────────
def sanitize_pdf_text(text):
    """Sanitize text for PDF generation by removing problematic characters."""
    if not isinstance(text, str):
        return str(text)
    
    # Replace problematic characters
    replacements = {
        '\x00': '',  # Null character
        '\r': ' ',   # Carriage return
        '\t': '    ', # Tab
    }
    
    result = text
    for old, new in replacements.items():
        result = result.replace(old, new)
    
    # Remove control characters except newline and basic whitespace
    result = ''.join(c if ord(c) >= 32 or c in '\n\t' else ' ' for c in result)
    
    return result


def generate_pdf_report(report_data):
    """Generate a professional PDF report using reportlab."""
    if not REPORTLAB_AVAILABLE:
        return None, "reportlab not installed. Run: pip install reportlab"

    try:
        import io
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch, cm
        from reportlab.lib import colors
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable, PageBreak
        from reportlab.lib.enums import TA_CENTER, TA_LEFT
        from reportlab.graphics.shapes import Drawing, Rect, String as RLString
        from reportlab.graphics.charts.barcharts import VerticalBarChart
    except ImportError as e:
        return None, f"Failed to import reportlab modules: {str(e)}"

    try:
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            rightMargin=1.5*cm,
            leftMargin=1.5*cm,
            topMargin=2*cm,
            bottomMargin=2*cm
        )
    except Exception as e:
        return None, f"Failed to create PDF document: {str(e)}"

    # Color palette
    DARK_BG    = colors.HexColor("#0f172a")
    TEAL       = colors.HexColor("#00e5ff")
    PURPLE     = colors.HexColor("#8b5cf6")
    GREEN      = colors.HexColor("#22c55e")
    RED        = colors.HexColor("#ef4444")
    AMBER      = colors.HexColor("#f59e0b")
    LIGHT_GRAY = colors.HexColor("#e2e8f0")
    MID_GRAY   = colors.HexColor("#94a3b8")
    WHITE      = colors.white

    styles = getSampleStyleSheet()

    title_style = ParagraphStyle(
        "TitleStyle",
        parent=styles["Title"],
        fontSize=22,
        textColor=DARK_BG,
        spaceAfter=6,
        fontName="Helvetica-Bold",
        alignment=TA_CENTER
    )
    subtitle_style = ParagraphStyle(
        "SubtitleStyle",
        parent=styles["Normal"],
        fontSize=11,
        textColor=MID_GRAY,
        spaceAfter=4,
        alignment=TA_CENTER
    )
    section_style = ParagraphStyle(
        "SectionStyle",
        parent=styles["Heading2"],
        fontSize=13,
        textColor=DARK_BG,
        spaceBefore=16,
        spaceAfter=6,
        fontName="Helvetica-Bold",
        borderPad=4,
    )
    body_style = ParagraphStyle(
        "BodyStyle",
        parent=styles["Normal"],
        fontSize=10,
        textColor=colors.HexColor("#1e293b"),
        spaceAfter=4,
        leading=15,
    )
    alert_high_style = ParagraphStyle(
        "AlertHigh",
        parent=styles["Normal"],
        fontSize=9,
        textColor=colors.HexColor("#7f1d1d"),
        backColor=colors.HexColor("#fef2f2"),
        borderColor=RED,
        borderWidth=1,
        borderPad=6,
        spaceAfter=6,
        leading=14,
    )
    alert_medium_style = ParagraphStyle(
        "AlertMedium",
        parent=styles["Normal"],
        fontSize=9,
        textColor=colors.HexColor("#78350f"),
        backColor=colors.HexColor("#fffbeb"),
        borderColor=AMBER,
        borderWidth=1,
        borderPad=6,
        spaceAfter=6,
        leading=14,
    )

    story = []
    metrics = report_data.get("metrics", {})
    mode = report_data.get("mode", "pre-training")
    filename = report_data.get("filename", "Unknown")
    created_at = report_data.get("created_at", "")
    bias_score = report_data.get("bias_score", metrics.get("bias_score", 0))
    risk_level = report_data.get("risk_level", metrics.get("risk_level", "UNKNOWN"))
    report_text = report_data.get("report", "")
    alert_explanations = report_data.get("alert_explanations", [])
    graphs = report_data.get("graphs", [])

    # Sanitize text for PDF generation
    filename = sanitize_pdf_text(filename)
    mode = sanitize_pdf_text(mode)
    created_at = sanitize_pdf_text(created_at)
    risk_level = sanitize_pdf_text(risk_level)
    report_text = sanitize_pdf_text(report_text)

    # ── Header ─────────────────────────────────────────────
    story.append(Paragraph("🔍 AI Fairness Auditor", title_style))
    story.append(Paragraph("Bias Detection & Fairness Report", subtitle_style))
    story.append(HRFlowable(width="100%", thickness=2, color=TEAL, spaceAfter=12))

    # ── Meta table ─────────────────────────────────────────
    risk_color = RED if risk_level in ("CRITICAL", "HIGH") else AMBER if risk_level == "MEDIUM" else GREEN
    score_color = RED if bias_score >= 60 else AMBER if bias_score >= 35 else GREEN

    meta_data = [
        ["File", sanitize_pdf_text(filename), "Mode", sanitize_pdf_text(mode.replace("-", " ").title())],
        ["Date", (sanitize_pdf_text(created_at[:10]) if created_at else "—"), "Bias Score", f"{bias_score}/100"],
        ["Risk Level", sanitize_pdf_text(risk_level), "Confidence", sanitize_pdf_text(str(metrics.get("audit_confidence", "—")))],
    ]
    meta_table = Table(meta_data, colWidths=[3*cm, 6*cm, 3*cm, 6*cm])
    meta_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#f8fafc")),
        ("BACKGROUND", (2, 0), (2, -1), colors.HexColor("#f8fafc")),
        ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
        ("FONTNAME", (2, 0), (2, -1), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("GRID", (0, 0), (-1, -1), 0.5, LIGHT_GRAY),
        ("ROWBACKGROUNDS", (0, 0), (-1, -1), [WHITE, colors.HexColor("#f8fafc")]),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ("LEFTPADDING", (0, 0), (-1, -1), 8),
    ]))
    story.append(meta_table)
    story.append(Spacer(1, 12))

    # ── Bias Score visual bar ──────────────────────────────
    story.append(Paragraph("Bias Risk Score", section_style))
    score_pct = min(bias_score / 100.0, 1.0)
    bar_width = 400
    bar_height = 20
    d = Drawing(bar_width, bar_height + 10)
    # Background bar
    bg = Rect(0, 5, bar_width, bar_height, fillColor=LIGHT_GRAY, strokeColor=None)
    d.add(bg)
    # Score fill
    fill = Rect(0, 5, bar_width * score_pct, bar_height, fillColor=score_color, strokeColor=None)
    d.add(fill)
    # Score label
    lbl = RLString(bar_width * score_pct + 4, 10, f"{bias_score}/100 — {risk_level}")
    lbl.fontSize = 9
    lbl.fillColor = colors.black
    d.add(lbl)
    story.append(d)
    story.append(Spacer(1, 12))

    # ── Key Metrics table ─────────────────────────────────
    story.append(Paragraph("Key Metrics", section_style))
    km_rows = [["Metric", "Value"]]
    if metrics.get("total_rows"):
        km_rows.append(["Total Rows", str(metrics["total_rows"])])
    if metrics.get("total_predictions"):
        km_rows.append(["Total Predictions", str(metrics["total_predictions"])])
    if metrics.get("total_columns"):
        km_rows.append(["Columns", str(metrics["total_columns"])])
    if metrics.get("protected_attributes"):
        attrs = ", ".join([sanitize_pdf_text(str(a)) for a in metrics["protected_attributes"]])
        km_rows.append(["Protected Attributes", attrs])
    if metrics.get("alert_count") is not None:
        km_rows.append(["Alerts Detected", str(metrics["alert_count"])])
    spd = metrics.get("statistical_parity_difference", {})
    if spd and isinstance(spd, dict):
        for attr, val in spd.items():
            if val is not None:
                km_rows.append([f"SPD ({sanitize_pdf_text(str(attr))})", f"{val:.4f}"])
    di = metrics.get("disparate_impact", {})
    if di and isinstance(di, dict):
        for attr, val in di.items():
            if val is not None:
                rule = "✓ PASS" if val >= 0.8 else "✗ FAIL"
                km_rows.append([f"Disparate Impact ({sanitize_pdf_text(str(attr))})", f"{val:.4f} — 80% Rule {rule}"])

    if len(km_rows) > 1:
        km_table = Table(km_rows, colWidths=[8*cm, 10*cm])
        km_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), DARK_BG),
            ("TEXTCOLOR", (0, 0), (-1, 0), WHITE),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("GRID", (0, 0), (-1, -1), 0.5, LIGHT_GRAY),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [WHITE, colors.HexColor("#f8fafc")]),
            ("TOPPADDING", (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ("LEFTPADDING", (0, 0), (-1, -1), 8),
        ]))
        story.append(km_table)
        story.append(Spacer(1, 12))

    # ── Graphs as bar charts ───────────────────────────────
    if graphs:
        story.append(Paragraph("Visual Analysis", section_style))
        for graph in graphs[:6]:  # Max 6 charts
            if not isinstance(graph, dict):
                continue
            labels = graph.get("labels", [])
            data = graph.get("data", graph.get("values", []))
            title = sanitize_pdf_text(graph.get("title", "Chart"))
            gtype = graph.get("type", "bar")

            if not labels or not data or len(labels) != len(data):
                continue

            try:
                float_data = [float(v) for v in data]
            except (TypeError, ValueError):
                continue

            if gtype == "pie":
                # Simple table for pie data
                story.append(Paragraph(title, body_style))
                pie_rows = [["Group", "Value"]] + [[sanitize_pdf_text(str(l)), str(v)] for l, v in zip(labels, float_data)]
                pie_table = Table(pie_rows, colWidths=[9*cm, 9*cm])
                pie_table.setStyle(TableStyle([
                    ("BACKGROUND", (0, 0), (-1, 0), PURPLE),
                    ("TEXTCOLOR", (0, 0), (-1, 0), WHITE),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, -1), 9),
                    ("GRID", (0, 0), (-1, -1), 0.5, LIGHT_GRAY),
                    ("TOPPADDING", (0, 0), (-1, -1), 4),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                ]))
                story.append(pie_table)
                story.append(Spacer(1, 8))
            else:
                # Bar chart using reportlab
                chart_height = 120
                chart_width = 400
                d = Drawing(chart_width, chart_height + 40)
                bc = VerticalBarChart()
                bc.x = 40
                bc.y = 20
                bc.width = chart_width - 60
                bc.height = chart_height - 10
                bc.data = [float_data]
                bc.categoryAxis.categoryNames = [sanitize_pdf_text(str(l)[:12]) for l in labels]
                bc.categoryAxis.labels.angle = 20 if len(labels) > 4 else 0
                bc.categoryAxis.labels.fontSize = 7
                bc.valueAxis.labels.fontSize = 7
                bc.bars[0].fillColor = TEAL
                bc.bars[0].strokeColor = None
                if float_data:
                    max_v = max(float_data)
                    bc.valueAxis.valueMax = max_v * 1.2 if max_v > 0 else 1
                    bc.valueAxis.valueMin = 0
                d.add(bc)

                title_lbl = RLString(chart_width / 2, chart_height + 25, title)
                title_lbl.fontSize = 9
                title_lbl.fontName = "Helvetica-Bold"
                title_lbl.fillColor = colors.black
                title_lbl.textAnchor = "middle"
                d.add(title_lbl)
                story.append(d)
                story.append(Spacer(1, 8))

    # ── Alert Explanations ────────────────────────────────
    if alert_explanations:
        story.append(Paragraph("Bias Detection Alerts", section_style))
        for i, alert in enumerate(alert_explanations):
            sev = sanitize_pdf_text((alert.get("severity") or "HIGH").upper())
            col_name = sanitize_pdf_text(alert.get("column", "Unknown"))
            explanation = sanitize_pdf_text(alert.get("explanation", ""))
            alert_style = alert_high_style if sev == "HIGH" or sev == "CRITICAL" else alert_medium_style
            header = f"⚠ Alert {i+1}: {col_name} — {sev} SEVERITY"
            story.append(Paragraph(header, alert_style))
            if explanation:
                # Clean up the explanation text
                clean_exp = explanation.replace("\n\n", " | ").replace("\n", " ")
                story.append(Paragraph(clean_exp[:500], body_style))
            story.append(Spacer(1, 4))

    # ── AI Report Text ────────────────────────────────────
    if report_text:
        story.append(PageBreak())
        story.append(Paragraph("AI-Generated Audit Report", section_style))
        story.append(HRFlowable(width="100%", thickness=1, color=TEAL, spaceAfter=8))

        # Parse into sections
        lines = report_text.split("\n")
        current_section = None
        current_body = []

        for line in lines:
            stripped = sanitize_pdf_text(line.strip())
            if not stripped:
                if current_body:
                    current_body.append("")
                continue

            is_heading = (
                (stripped[0].isdigit() and ". " in stripped[:4]) or
                (stripped == stripped.upper() and len(stripped) > 4 and len(stripped) < 60)
            )

            if is_heading:
                if current_section and current_body:
                    story.append(Paragraph(current_section, section_style))
                    for bl in current_body:
                        if bl:
                            story.append(Paragraph(bl, body_style))
                    current_body = []
                current_section = stripped.lstrip("0123456789. ")
            else:
                current_body.append(stripped)

        if current_section and current_body:
            story.append(Paragraph(current_section, section_style))
            for bl in current_body:
                if bl:
                    story.append(Paragraph(bl, body_style))

    # ── Footer note ───────────────────────────────────────
    story.append(Spacer(1, 20))
    story.append(HRFlowable(width="100%", thickness=1, color=LIGHT_GRAY))
    story.append(Paragraph(
        f"Generated by AI Fairness Auditor · {datetime.now().strftime('%Y-%m-%d %H:%M')} · Confidential",
        ParagraphStyle("Footer", parent=styles["Normal"], fontSize=8, textColor=MID_GRAY, alignment=TA_CENTER)
    ))

    try:
        doc.build(story)
        buffer.seek(0)
        return buffer.getvalue(), None
    except Exception as e:
        error_msg = f"PDF generation error: {str(e)}"
        return None, error_msg


# ─── Fix Dataset ──────────────────────────────────────────────────────────────
def _repair_target_binary(df, target_col):
    if not target_col or target_col not in df.columns:
        return None
    return convert_to_binary(df[target_col])


def _balance_representation(df, attr, target_share=0.45, random_state=42):
    """Oversample underrepresented groups toward a fair representation share."""
    if attr not in df.columns:
        return df, False

    working = df.copy()
    counts = working[attr].astype(str).value_counts()
    if len(counts) < 2:
        return working, False

    max_share = float(counts.max() / len(working))
    if max_share <= 0.62:
        return working, False

    n_groups = len(counts)
    ideal_count = max(10, int(len(working) * target_share))
    frames = [working]
    changed = False

    for group_name, group_count in counts.items():
        if group_count >= ideal_count:
            continue
        needed = ideal_count - int(group_count)
        group_df = working[working[attr].astype(str) == str(group_name)]
        if len(group_df) == 0 or needed <= 0:
            continue

        target_col = detect_target_column(working)
        if target_col and target_col in working.columns:
            y = _repair_target_binary(group_df, target_col)
            pos_df = group_df[y == 1]
            neg_df = group_df[y == 0]
            pos_rate = len(pos_df) / max(len(group_df), 1)
            n_pos = max(0, min(needed, int(round(needed * pos_rate))))
            n_neg = needed - n_pos
            extras = []
            if n_pos > 0 and len(pos_df) > 0:
                extras.append(pos_df.sample(n=n_pos, replace=True, random_state=random_state))
            if n_neg > 0 and len(neg_df) > 0:
                extras.append(neg_df.sample(n=n_neg, replace=True, random_state=random_state))
            if extras:
                frames.extend(extras)
                changed = True
        else:
            frames.append(group_df.sample(n=needed, replace=True, random_state=random_state))
            changed = True

    if not changed:
        return working, False

    result = pd.concat(frames, ignore_index=True)
    return result.sample(frac=1, random_state=random_state).reset_index(drop=True), True


def _equalize_outcome_rates(df, attr, target_col, target_rate=None, random_state=42):
    """
    Align per-group positive rates using outcome-aware resampling.
    Disadvantaged groups get positive upsampling; advantaged groups get negative upsampling.
    """
    if attr not in df.columns or not target_col or target_col not in df.columns:
        return df, False

    working = df.copy()
    y = _repair_target_binary(working, target_col)
    working["_repair_y_"] = y

    group_stats = working.groupby(working[attr].astype(str))["_repair_y_"].agg(["sum", "count", "mean"])
    if len(group_stats) < 2:
        working.drop(columns=["_repair_y_"], inplace=True, errors="ignore")
        return working, False

    rates = group_stats["mean"].astype(float)
    max_rate = float(rates.max())
    min_rate = float(rates.min())
    if max_rate - min_rate <= 0.04:
        working.drop(columns=["_repair_y_"], inplace=True, errors="ignore")
        return working, False

    if target_rate is None:
        # Pull all groups toward the median positive rate for fairer parity.
        target_rate = float(rates.median())
        target_rate = min(max(target_rate, min_rate + 0.02), max_rate - 0.02)

    frames = [working]
    changed = False

    for group_name, row in group_stats.iterrows():
        count = int(row["count"])
        pos = int(row["sum"])
        rate = float(row["mean"])
        if count == 0:
            continue

        group_df = working[working[attr].astype(str) == str(group_name)]
        pos_df = group_df[group_df["_repair_y_"] == 1]
        neg_df = group_df[group_df["_repair_y_"] == 0]

        if rate < target_rate - 0.015:
            desired_pos = max(pos + 1, int(round(target_rate * count)))
            extra_pos = desired_pos - pos
            if extra_pos > 0 and len(pos_df) > 0:
                frames.append(pos_df.sample(n=extra_pos, replace=True, random_state=random_state))
                changed = True
        elif rate > target_rate + 0.015:
            # Add negative examples to lower the observed positive rate without deleting rows.
            if rate > 0:
                desired_total = max(count + 1, int(round(pos / target_rate)))
                extra_neg = desired_total - count
                if extra_neg > 0 and len(neg_df) > 0:
                    frames.append(neg_df.sample(n=extra_neg, replace=True, random_state=random_state))
                    changed = True

    working.drop(columns=["_repair_y_"], inplace=True, errors="ignore")
    if not changed:
        return df, False

    result = pd.concat(frames, ignore_index=True)
    return result.sample(frac=1, random_state=random_state).reset_index(drop=True), True


def _kamiran_calders_resample(df, attr, target_col, random_state=42):
    """
    Reweighing-inspired repair: resample to balance joint (protected attribute, outcome) cells.
    """
    if attr not in df.columns or not target_col or target_col not in df.columns:
        return df, False

    working = df.copy()
    y = _repair_target_binary(working, target_col)
    working["_repair_y_"] = y
    cell_counts = working.groupby([working[attr].astype(str), working["_repair_y_"]]).size()
    if len(cell_counts) < 4:
        working.drop(columns=["_repair_y_"], inplace=True, errors="ignore")
        return working, False

    total = len(working)
    n_groups = working[attr].astype(str).nunique()
    n_outcomes = 2
    target_cell = max(5, int(total / (n_groups * n_outcomes)))

    frames = [working]
    changed = False
    for (group_name, outcome), count in cell_counts.items():
        if int(count) >= target_cell:
            continue
        needed = target_cell - int(count)
        cell_df = working[
            (working[attr].astype(str) == str(group_name)) & (working["_repair_y_"] == outcome)
        ]
        if len(cell_df) == 0 or needed <= 0:
            continue
        frames.append(cell_df.sample(n=needed, replace=True, random_state=random_state))
        changed = True

    working.drop(columns=["_repair_y_"], inplace=True, errors="ignore")
    if not changed:
        return df, False

    result = pd.concat(frames, ignore_index=True)
    return result.sample(frac=1, random_state=random_state).reset_index(drop=True), True


def apply_fairness_repair(df, metrics=None, aggressive=False):
    """
    Multi-stage fairness repair:
    1) Kamiran-Calders joint resampling
    2) Representation balancing
    3) Iterative outcome-parity equalization on worst SPD attributes
    """
    working = _coerce_dataframe_types(df.copy())
    original_len = len(working)
    if original_len < 5:
        return working, "Dataset too small to repair."

    if metrics is None:
        metrics = compute_pretraining_metrics(working)

    protected_attrs = metrics.get("protected_attributes") or detect_protected_attributes(bin_age_column(working))
    target_col = metrics.get("target_column") or detect_target_column(working)
    if not protected_attrs:
        return working, "No protected attributes detected; dataset unchanged."

    random_state = 42
    max_rows = int(original_len * (3.0 if aggressive else 2.5))
    steps = []
    share_target = 0.40 if aggressive else 0.45

    # Stage 1: joint (attribute, outcome) balancing when labels exist
    if target_col and target_col in working.columns:
        for attr in protected_attrs:
            working, changed = _kamiran_calders_resample(working, attr, target_col, random_state)
            if changed:
                steps.append(f"Joint outcome-representation balancing on '{attr}'")
                if len(working) > max_rows:
                    working = working.sample(n=max_rows, random_state=random_state).reset_index(drop=True)

    # Stage 2: representation repair
    for attr in protected_attrs:
        working, changed = _balance_representation(working, attr, target_share=share_target, random_state=random_state)
        if changed:
            steps.append(f"Representation rebalancing on '{attr}'")
            if len(working) > max_rows:
                working = working.sample(n=max_rows, random_state=random_state).reset_index(drop=True)

    # Stage 3: iterative outcome parity on worst attributes
    if target_col and target_col in working.columns:
        passes = 4 if aggressive else 3
        for pass_idx in range(passes):
            current_metrics = compute_pretraining_metrics(working)
            spd_values = current_metrics.get("statistical_parity_difference", {})
            if not spd_values:
                break

            worst_attr = max(spd_values, key=lambda a: float(spd_values.get(a) or 0))
            worst_spd = float(spd_values.get(worst_attr) or 0)
            if worst_spd <= 0.04:
                break

            rates = current_metrics.get("outcome_bias", {}).get(worst_attr, {})
            target_rate = None
            if rates:
                values = [float(v) for v in rates.values()]
                target_rate = float(np.median(values))
                if aggressive:
                    target_rate = float(np.mean(values))

            working, changed = _equalize_outcome_rates(
                working, worst_attr, target_col, target_rate=target_rate, random_state=random_state + pass_idx
            )
            if changed:
                steps.append(f"Outcome parity pass {pass_idx + 1} on '{worst_attr}' (SPD {worst_spd:.3f})")
                if len(working) > max_rows:
                    working = working.sample(n=max_rows, random_state=random_state).reset_index(drop=True)
            else:
                break

    working = working.sample(frac=1, random_state=random_state).reset_index(drop=True)

    if not steps:
        explanation = "Dataset already near fairness thresholds; light balancing applied."
    else:
        explanation = (
            "Advanced multi-stage fairness repair applied: "
            + "; ".join(steps[:6])
            + ("." if len(steps) <= 6 else "; additional refinement passes.")
        )
    return working, explanation


def _diagnose_dataset(df, metrics):
    """
    Pure-Python diagnosis of what kind of bias exists and what strategy to use.
    Returns a dict with diagnosis text, chosen strategy, and parameters.
    """
    protected_attrs = metrics.get("protected_attributes", [])
    spd_values      = metrics.get("statistical_parity_difference", {})
    di_values       = metrics.get("disparate_impact", {})
    distributions   = metrics.get("distributions", {})
    bias_score      = metrics.get("bias_score", 0)
    alerts          = metrics.get("alerts", [])

    issues = []
    strategy_parts = []
    fix_params = {}

    # 1. Check distribution imbalance per attribute
    for attr, dist in distributions.items():
        if not dist:
            continue
        max_val = max(dist.values())
        dominant = max(dist, key=dist.get)
        minority_groups = {g: v for g, v in dist.items() if v < 0.25}
        if max_val > 0.70:
            issues.append(
                f"'{attr}' is severely imbalanced: '{dominant}' dominates at {round(max_val*100,1)}%. "
                f"Minority groups {list(minority_groups.keys())} are underrepresented."
            )
            strategy_parts.append("outcome-aware oversampling")
            fix_params.setdefault("oversample_attrs", {})[attr] = {
                "dominant": dominant,
                "minority": list(minority_groups.keys()),
                "target_pct": min(0.6, max_val)
            }

    # 2. Check outcome bias (SPD / DI)
    worst_spd_attr = None
    worst_spd_val  = 0
    for attr, spd in spd_values.items():
        if spd and float(spd) > worst_spd_val:
            worst_spd_val  = float(spd)
            worst_spd_attr = attr

    if worst_spd_attr and worst_spd_val > 0.10:
        di = float(di_values.get(worst_spd_attr, 1.0) or 1.0)
        rule80 = "fails" if di < 0.80 else "passes"
        issues.append(
            f"Outcome bias in '{worst_spd_attr}': SPD={round(worst_spd_val,4)}, DI={round(di,4)} "
            f"({rule80} the 80% fairness rule). One group is {round(worst_spd_val*100,1)}% "
            f"more likely to receive a positive outcome."
        )
        strategy_parts.append("outcome-rate rebalancing via upsampling minority positive outcomes")
        fix_params["outcome_attr"] = worst_spd_attr
        fix_params["outcome_spd"]  = worst_spd_val

    # 3. Build human-readable diagnosis
    if not issues:
        diagnosis = (
            f"Dataset bias score is {bias_score}/100. No severe imbalance or outcome bias detected "
            f"above threshold. Applying gentle rebalancing as a precaution."
        )
        strategy_parts = ["gentle proportional rebalancing"]
    else:
        diagnosis = (
            f"Bias score: {bias_score}/100. {len(issues)} issue(s) detected across "
            f"{len(protected_attrs)} protected attribute(s). "
            + " | ".join(issues)
        )

    strategy = " + ".join(strategy_parts) if strategy_parts else "proportional rebalancing"
    return {"diagnosis": diagnosis, "strategy": strategy, "fix_params": fix_params, "issues": issues}


def _build_fix_code(df, metrics, diagnosis_result):
    """
    Build a deterministic, correct Python fix function based on diagnosis.
    This is the PRIMARY fix — no AI needed. AI is only used for explanation.
    """
    protected_attrs = metrics.get("protected_attributes", [])
    distributions   = metrics.get("distributions", {})
    spd_values      = metrics.get("statistical_parity_difference", {})
    fix_params      = diagnosis_result.get("fix_params", {})
    target_col      = metrics.get("target_column")

    oversample_config = json.dumps(fix_params.get("oversample_attrs", {}))
    outcome_attr      = fix_params.get("outcome_attr", "")
    outcome_spd       = fix_params.get("outcome_spd", 0)
    target_col_safe   = json.dumps(target_col) if target_col else "None"
    protected_safe    = json.dumps(protected_attrs)

    code = f'''
def fix_dataset(df):
    import pandas as pd
    import numpy as np

    df = df.copy()
    # Convert numeric columns safely
    for col in df.columns:
        try:
            df[col] = pd.to_numeric(df[col])
        except Exception:
            pass

    protected_attrs  = {protected_safe}
    oversample_cfg   = {oversample_config}
    outcome_attr     = {json.dumps(outcome_attr)}
    target_col       = {target_col_safe}
    random_state     = 42
    np.random.seed(random_state)

    frames = [df]

    # ── STEP 1: Outcome-aware oversampling ─────────────────────────────
    # For each imbalanced protected attribute, oversample minority groups
    # BUT preserve the SAME outcome ratio as the majority group to avoid
    # inflating bias in the opposite direction.
    for attr, cfg in oversample_cfg.items():
        if attr not in df.columns:
            continue
        dominant   = cfg.get("dominant", "")
        minority   = cfg.get("minority", [])
        if not minority:
            continue

        # Count majority group rows
        majority_df = df[df[attr].astype(str) == str(dominant)]
        majority_count = max(len(majority_df), 10)
        target_count   = max(10, int(majority_count * 0.55))  # target 55% of majority

        for group in minority:
            group_df = df[df[attr].astype(str) == str(group)]
            if len(group_df) == 0:
                continue
            current = len(group_df)
            if current >= target_count:
                continue
            needed = target_count - current

            if target_col and target_col in df.columns:
                # Stratified oversample: preserve outcome ratio
                pos_df  = group_df[group_df[target_col].astype(str).isin(
                    ["1","yes","true","approved","accept","selected","hired","positive"])]
                neg_df  = group_df[~group_df[target_col].astype(str).isin(
                    ["1","yes","true","approved","accept","selected","hired","positive"])]
                pos_r = len(pos_df) / max(len(group_df), 1)
                n_pos = max(0, int(needed * pos_r))
                n_neg = needed - n_pos
                extra = []
                if n_pos > 0 and len(pos_df) > 0:
                    extra.append(pos_df.sample(n=n_pos, replace=True, random_state=random_state))
                if n_neg > 0 and len(neg_df) > 0:
                    extra.append(neg_df.sample(n=n_neg, replace=True, random_state=random_state))
                if extra:
                    frames.extend(extra)
            else:
                # No target col — just oversample rows directly
                extra = group_df.sample(n=needed, replace=True, random_state=random_state)
                frames.append(extra)

    # ── STEP 2: Outcome-rate rebalancing ───────────────────────────────
    # If outcome bias detected (SPD > 0.10), upsample the positive outcomes
    # of disadvantaged groups so their positive rate approaches the average.
    if outcome_attr and target_col and outcome_attr in df.columns and target_col in df.columns:
        merged = pd.concat(frames, ignore_index=True) if len(frames) > 1 else df.copy()

        POSITIVE = {{"1","yes","true","approved","accept","selected","hired","positive"}}
        merged["_pos_"] = merged[target_col].astype(str).str.lower().isin(POSITIVE).astype(int)

        group_rates = merged.groupby(merged[outcome_attr].astype(str))["_pos_"].mean()
        if len(group_rates) >= 2:
            avg_rate = float(group_rates.mean())
            for group_name, rate in group_rates.items():
                if float(rate) < avg_rate - 0.05:
                    # Upsample positive outcomes for this group
                    g_df  = merged[merged[outcome_attr].astype(str) == group_name]
                    pos   = g_df[g_df["_pos_"] == 1]
                    if len(pos) == 0:
                        continue
                    neg   = g_df[g_df["_pos_"] == 0]
                    target_pos = int(len(g_df) * min(avg_rate, 0.70))
                    extra_pos  = max(0, target_pos - len(pos))
                    if extra_pos > 0:
                        frames.append(pos.sample(n=extra_pos, replace=True, random_state=random_state))

        merged.drop(columns=["_pos_"], inplace=True, errors="ignore")

    # ── STEP 3: Combine and shuffle ────────────────────────────────────
    if len(frames) > 1:
        result = pd.concat(frames, ignore_index=True)
    else:
        result = df.copy()

    # Safety: never return more than 3x the original rows
    max_rows = len(df) * 3
    if len(result) > max_rows:
        result = result.sample(n=max_rows, random_state=random_state)

    result = result.sample(frac=1, random_state=random_state).reset_index(drop=True)

    # Drop temp columns if any leaked
    result.drop(columns=["_pos_"], inplace=True, errors="ignore")

    return result
'''
    return code


@app.route("/api/fix_dataset", methods=["POST"])
@login_required
def api_fix_dataset():
    data = request.get_json()
    if not data:
        return jsonify({"error": "No data sent"}), 400

    metrics = data.get("metrics", {})
    df_data = data.get("df_data", [])

    if not df_data:
        return jsonify({"error": "No dataset provided. Please upload a CSV file — text/PDF files cannot be repaired."}), 400

    try:
        df = _coerce_dataframe_types(pd.DataFrame(df_data))
    except Exception as e:
        return jsonify({"error": f"Invalid dataset: {str(e)}"}), 400

    if len(df) < 5:
        return jsonify({"error": "Dataset too small for repair (need at least 5 rows)."}), 400

    protected_attrs = metrics.get("protected_attributes", [])
    bias_score = metrics.get("bias_score", 0)

    # ── Step 1: Run advanced deterministic fairness repair ─────────────────
    before_metrics = compute_pretraining_metrics(df)
    before_score = before_metrics.get("bias_score", bias_score)

    fixed_df, algorithm_explanation = apply_fairness_repair(df, before_metrics, aggressive=False)
    after_metrics = compute_pretraining_metrics(fixed_df)
    after_score = after_metrics.get("bias_score", 0)

    def _fails_thresholds(metric_snapshot):
        for spd_val in metric_snapshot.get("statistical_parity_difference", {}).values():
            if spd_val is not None and float(spd_val) > 0.10:
                return True
        for di_val in metric_snapshot.get("disparate_impact", {}).values():
            if di_val is not None and float(di_val) < 0.80:
                return True
        return False

    # ── Step 2: If still biased, run a more aggressive second pass ─────────
    if after_score >= before_score or _fails_thresholds(after_metrics):
        aggressive_df, aggressive_expl = apply_fairness_repair(df, before_metrics, aggressive=True)
        aggressive_metrics = compute_pretraining_metrics(aggressive_df)
        aggressive_score = aggressive_metrics.get("bias_score", 0)

        if aggressive_score < after_score or (
            aggressive_score <= after_score and not _fails_thresholds(aggressive_metrics)
        ):
            fixed_df = aggressive_df
            after_metrics = aggressive_metrics
            after_score = aggressive_score
            algorithm_explanation = aggressive_expl + " (aggressive refinement pass)"

    if fixed_df is None or len(fixed_df) == 0:
        return jsonify({"error": "Dataset repair produced empty result."}), 500

    improvement = max(0, round(before_score - after_score, 1))

    # ── Step 5: Convert fixed dataset to CSV ──────────────────────────────
    fixed_csv = fixed_df.to_csv(index=False)

    # ── Step 6: Save to MongoDB Reports ──────────────────────────────────
    report_id = None
    if MONGO_AVAILABLE:
        try:
            doc = normalize_for_mongo({
                "user_id": session["user_id"],
                "mode": "fix_dataset",
                "filename": f"Fixed Dataset - {improvement}% improvement",
                "file_name": f"Fixed Dataset - {improvement}% improvement",
                "algorithm_explanation": algorithm_explanation,
                "before_score": before_score,
                "after_score": after_score,
                "bias_score": after_score,
                "improvement": improvement,
                "before_metrics": before_metrics,
                "after_metrics": after_metrics,
                "metrics": after_metrics,
                "original_rows": len(df),
                "fixed_rows": len(fixed_df),
                "protected_attributes": protected_attrs,
                "fixed_csv": fixed_csv,
                "risk_level": "LOW" if after_score < 35 else "MEDIUM" if after_score < 60 else "HIGH",
                "report": f"Dataset fix simulation completed.\n\nBefore Score: {before_score}/100\nAfter Score: {after_score}/100\nImprovement: {improvement}%\n\n{algorithm_explanation}",
                "created_at": datetime.now(timezone.utc).isoformat()
            })
            inserted = db["reports"].insert_one(doc)
            report_id = str(inserted.inserted_id)
        except Exception as e:
            print(f"[Fix Dataset] MongoDB error: {e}")
            pass

    passes = True
    for spd_val in after_metrics.get("statistical_parity_difference", {}).values():
        if abs(spd_val) > 0.10:
            passes = False
    for di_val in after_metrics.get("disparate_impact", {}).values():
        if di_val < 0.80:
            passes = False

    return jsonify(normalize_for_mongo({
        "success": True,
        "report_id": report_id,
        "algorithm_explanation": algorithm_explanation,
        "before_metrics": before_metrics,
        "after_metrics": after_metrics,
        "bias_score_improvement": improvement,
        "original_rows": len(df),
        "fixed_rows": len(fixed_df),
        "passes": passes,
        "fixed_csv": fixed_csv,
    }))


def _get_fallback_fix_code(metrics):
    """Returns a reliable fallback fix function when AI is unavailable."""
    protected_attrs = metrics.get("protected_attributes", [])
    distributions = metrics.get("distributions", {})
    target_column = metrics.get("target_column")
    attrs_str = json.dumps(protected_attrs)
    dist_str = json.dumps(distributions)
    target_col_safe = json.dumps(target_column) if target_column else "None"

    return f"""
def fix_dataset(df):
    import pandas as pd
    import numpy as np

    df = df.copy()
    protected_attrs = {attrs_str}
    distributions = {dist_str}
    target_col = {target_col_safe}

    def to_binary(series):
        s = series.astype(str).str.lower().str.strip()
        return s.isin(['1','yes','true','approved','accept','accepted','selected','hired','positive','pass']).astype(int)

    if target_col is not None and target_col in df.columns:
        df['_target_bin'] = to_binary(df[target_col])

    for attr in protected_attrs:
        if attr not in df.columns:
            continue
        dist = distributions.get(attr, {{}})
        if not dist:
            continue

        majority_group = max(dist, key=dist.get)
        minority_groups = [g for g, v in dist.items() if v < 0.35]
        majority_count = len(df[df[attr].astype(str) == str(majority_group)])
        target_count = max(10, int(majority_count * 0.65))

        frames = [df]
        for group in minority_groups:
            group_df = df[df[attr].astype(str) == str(group)]
            if len(group_df) == 0:
                continue
            current_count = len(group_df)
            if current_count >= target_count:
                continue
            needed = target_count - current_count

            if target_col is not None and '_target_bin' in df.columns:
                pos_df = group_df[group_df['_target_bin'] == 1]
                neg_df = group_df[group_df['_target_bin'] == 0]
                pos_ratio = len(pos_df) / max(len(group_df), 1)
                target_pos = min(len(group_df) + needed, int(max(pos_ratio, 0.45) * (len(group_df) + needed)))
                n_pos = max(0, min(needed, target_pos - len(pos_df)))
                n_neg = needed - n_pos
                extra = []
                if n_pos > 0 and len(pos_df) > 0:
                    extra.append(pos_df.sample(n=n_pos, replace=True, random_state=42))
                if n_neg > 0 and len(neg_df) > 0:
                    extra.append(neg_df.sample(n=n_neg, replace=True, random_state=42))
                if extra:
                    frames.extend(extra)
            else:
                frames.append(group_df.sample(n=needed, replace=True, random_state=42))

        if len(frames) > 1:
            df = pd.concat(frames, ignore_index=True)
            df = df.sample(frac=1, random_state=42).reset_index(drop=True)

    if target_col is not None and '_target_bin' in df.columns:
        df = df.drop(columns=['_target_bin'], errors='ignore')

    return df
"""


# ─── PDF Export Route ─────────────────────────────────────────────────────────
@app.route("/api/export_pdf", methods=["POST"])
@login_required
def api_export_pdf():
    data = request.get_json()
    if not data:
        return jsonify({"error": "No data sent"}), 400

    report_id = data.get("report_id")
    
    print(f"[PDF Export] report_id={report_id}, MONGO_AVAILABLE={MONGO_AVAILABLE}, REPORTLAB_AVAILABLE={REPORTLAB_AVAILABLE}")

    if not REPORTLAB_AVAILABLE:
        return jsonify({"error": "PDF generation requires reportlab. Run: pip install reportlab"}), 500

    # Try to fetch from MongoDB if available and report_id provided
    report_data = None
    if report_id and MONGO_AVAILABLE:
        try:
            report_data = db["reports"].find_one({"_id": ObjectId(report_id), "user_id": session["user_id"]})
            if not report_data:
                return jsonify({"error": "Report not found in database"}), 404
            report_data["_id"] = str(report_data["_id"])
            print(f"[PDF Export] Fetched report from MongoDB, keys: {list(report_data.keys())}")
        except Exception as e:
            err_msg = f"MongoDB error: {str(e)}"
            print(f"[PDF Export] {err_msg}")
            return jsonify({"error": err_msg}), 500
    
    # Otherwise, check if report data is provided directly in the request
    if not report_data:
        report_data = data.get("report_data")
        if not report_data:
            return jsonify({"error": "No report_id or report_data provided"}), 400
        print(f"[PDF Export] Using report_data from request, keys: {list(report_data.keys()) if isinstance(report_data, dict) else 'N/A'}")

    if not isinstance(report_data, dict):
        return jsonify({"error": "Invalid report data format"}), 400

    pdf_bytes, err = generate_pdf_report(report_data)
    if err:
        print(f"[PDF Export] PDF generation error: {err}")
        return jsonify({"error": err}), 500

    filename = report_data.get("filename", "report")
    safe_name = secure_filename(filename).replace(" ", "_")

    response = make_response(pdf_bytes)
    response.headers["Content-Type"] = "application/pdf"
    response.headers["Content-Disposition"] = f'attachment; filename="fairness-report-{safe_name}.pdf"'
    print(f"[PDF Export] PDF exported successfully")
    return response


# ─── Download Fixed Dataset CSV ───────────────────────────────────────────────
@app.route("/api/download_dataset/<report_id>", methods=["GET"])
@login_required
def download_dataset(report_id):
    """Download fixed dataset CSV from fix_dataset report."""
    if not report_id:
        return jsonify({"error": "Report ID missing"}), 400
    
    if not MONGO_AVAILABLE:
        return jsonify({"error": "Database unavailable"}), 503
    
    try:
        report_data = db["reports"].find_one({
            "_id": ObjectId(report_id), 
            "user_id": session["user_id"],
            "mode": "fix_dataset"
        })
        if not report_data:
            return jsonify({"error": "Dataset not found or access denied"}), 404
        
        fixed_csv = report_data.get("fixed_csv")
        if not fixed_csv:
            return jsonify({"error": "No dataset available in this report"}), 400
        
        filename = report_data.get("filename", "fixed_dataset")
        safe_name = secure_filename(filename).replace(" ", "_")
        
        response = make_response(fixed_csv)
        response.headers["Content-Type"] = "text/csv"
        response.headers["Content-Disposition"] = f'attachment; filename="{safe_name}.csv"'
        return response
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ─── Auth Routes ──────────────────────────────────────────────────────────────
@app.route("/")
def index():
    return render_template("index.html")

@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        email = request.form.get("email", "").strip().lower()
        password = request.form.get("password", "")
        if not MONGO_AVAILABLE:
            session["user_id"] = "demo_user"
            session["user_email"] = email
            session["user_name"] = email.split("@")[0]
            return redirect(url_for("dashboard"))
        user = db["users"].find_one({"email": email})
        if user and check_password_hash(user["password"], password):
            session["user_id"] = str(user["_id"])
            session["user_email"] = email
            session["user_name"] = user.get("name", email.split("@")[0])
            return redirect(url_for("dashboard"))
        flash("Invalid email or password", "error")
    return render_template("login.html")

@app.route("/register", methods=["GET", "POST"])
def register():
    if request.method == "POST":
        name = request.form.get("name", "").strip()
        email = request.form.get("email", "").strip().lower()
        password = request.form.get("password", "")
        if not MONGO_AVAILABLE:
            session["user_id"] = "demo_user"
            session["user_email"] = email
            session["user_name"] = name
            return redirect(url_for("dashboard"))
        if db["users"].find_one({"email": email}):
            flash("Email already registered", "error")
            return render_template("register.html")
        user_id = db["users"].insert_one({
            "name": name, "email": email,
            "password": generate_password_hash(password),
            "created_at": datetime.now(timezone.utc)
        }).inserted_id
        session["user_id"] = str(user_id)
        session["user_email"] = email
        session["user_name"] = name
        return redirect(url_for("dashboard"))
    return render_template("register.html")

@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("index"))

@app.route("/dashboard")
@login_required
def dashboard():
    provider = get_ai_provider()
    return render_template("dashboard.html",
                           user_name=session.get("user_name", "User"),
                           ai_provider=provider,
                           mongo_available=MONGO_AVAILABLE)


# ─── Pre-Training API ─────────────────────────────────────────────────────────
@app.route("/api/pretrain", methods=["POST"])
@login_required
def api_pretrain():
    if "file" not in request.files:
        return jsonify({"error": "No file uploaded"}), 400
    file = request.files["file"]
    if not file.filename:
        return jsonify({"error": "No file selected"}), 400

    mode, data = load_uploaded_file(file)
    if mode == "error":
        return jsonify({"error": data}), 400

    if mode == "text":
        filename_safe = secure_filename(file.filename) if file.filename else "document"
        metrics = compute_text_metrics(data, filename_safe)
        graphs = generate_text_graph_data(metrics)

        groq_explanations = []
        for g in graphs:
            explanation = explain_graph_with_groq({"title": g["title"], "labels": g["labels"], "data": g["data"]})
            groq_explanations.append({"chart": g["title"], "explanation": explanation})

        target_col = metrics.get("target_column")
        alert_explanations = []
        for alert in metrics.get("alerts", []):
            explanation = generate_alert_explanation_groq(alert, target_col)
            alert_explanations.append({
                "column": alert.get("column"),
                "type": alert.get("type"),
                "severity": alert.get("severity", "HIGH"),
                "explanation": explanation,
                "spd": alert.get("statistical_parity_difference"),
                "di": alert.get("disparate_impact"),
                "dominant_group": alert.get("dominant_group"),
                "pct": alert.get("pct"),
                "target": alert.get("target"),
            })

        report_result = generate_ai_report(metrics, "pre")
        result = {
            "mode": "text_analysis",
            "metrics": metrics,
            "graphs": graphs,
            "groq_explanations": groq_explanations,
            "ai_explanations": groq_explanations,
            "alert_explanations": alert_explanations,
            "report": report_result.get("report", ""),
            "report_error": report_result.get("error"),
            "provider": report_result.get("provider", get_ai_provider()),
            "filename": filename_safe,
            "message": f"Analysed as unstructured text ({metrics['file_type'].upper()}). "
                       f"Demographic mentions extracted: {metrics['text_stats']['demographic_mentions']}. "
                       f"Full bias metrics computed from co-occurrence analysis."
        }

        if MONGO_AVAILABLE:
            doc = normalize_for_mongo({
                "user_id": session["user_id"],
                "mode": "pre-training",
                "filename": filename_safe,
                "metrics": metrics,
                "graphs": graphs,
                "groq_explanations": groq_explanations,
                "ai_explanations": groq_explanations,
                "alert_explanations": alert_explanations,
                "report": report_result.get("report", ""),
                "provider": report_result.get("provider", get_ai_provider()),
                "bias_score": metrics.get("bias_score", 0),
                "risk_level": metrics.get("risk_level", "UNKNOWN"),
                "created_at": datetime.now(timezone.utc).isoformat()
            })
            inserted = db["reports"].insert_one(doc)
            result["report_id"] = str(inserted.inserted_id)

        return jsonify(normalize_for_mongo(result))

    df = data
    metrics = compute_pretraining_metrics(df)
    graphs = generate_graph_data(metrics, "pre")

    groq_explanations = []
    for g in graphs:
        explanation = explain_graph_with_groq({"title": g["title"], "labels": g["labels"], "data": g["data"]})
        groq_explanations.append({"chart": g["title"], "explanation": explanation})

    target_col = metrics.get("target_column")
    alert_explanations = []
    for alert in metrics.get("alerts", []):
        explanation = generate_alert_explanation_groq(alert, target_col)
        alert_explanations.append({
            "column": alert.get("column"),
            "type": alert.get("type"),
            "severity": alert.get("severity", "HIGH"),
            "explanation": explanation,
            "spd": alert.get("statistical_parity_difference"),
            "di": alert.get("disparate_impact"),
            "dominant_group": alert.get("dominant_group"),
            "pct": alert.get("pct"),
            "target": alert.get("target"),
        })

    report_result = generate_ai_report(metrics, "pre")
    result = {
        "metrics": metrics,
        "graphs": graphs,
        "groq_explanations": groq_explanations,
        "ai_explanations": groq_explanations,
        "alert_explanations": alert_explanations,
        "report": report_result.get("report", ""),
        "report_error": report_result.get("error"),
        "provider": report_result.get("provider", get_ai_provider()),
        "filename": secure_filename(file.filename) if file.filename else "unknown"
    }

    if MONGO_AVAILABLE:
        doc = normalize_for_mongo({
            "user_id": session["user_id"],
            "mode": "pre-training",
            "filename": result["filename"],
            "metrics": metrics,
            "graphs": graphs,
            "groq_explanations": groq_explanations,
            "ai_explanations": groq_explanations,
            "alert_explanations": alert_explanations,
            "report": report_result.get("report", ""),
            "provider": report_result.get("provider", get_ai_provider()),
            "bias_score": metrics.get("bias_score", 0),
            "risk_level": metrics.get("risk_level", "UNKNOWN"),
            "created_at": datetime.now(timezone.utc).isoformat()
        })
        inserted = db["reports"].insert_one(doc)
        result["report_id"] = str(inserted.inserted_id)

    return jsonify(normalize_for_mongo(result))


# ─── Post-Training API ────────────────────────────────────────────────────────
@app.route("/api/posttrain", methods=["POST"])
@login_required
def api_posttrain():
    df = None

    if "file" in request.files and request.files["file"].filename:
        mode, data = load_uploaded_file(request.files["file"])
        if mode == "error":
            return jsonify({"error": data}), 400
        if mode == "text":
            return jsonify({"error": "Please upload CSV/JSON/XLSX for post-training audit.", "switch_to_text": True}), 400
        df = data
        # Auto-detect prediction column — no user input needed
        prediction_col = detect_target_column(df)
        if not prediction_col:
            # Use last column as prediction column
            prediction_col = df.columns[-1]
        label_col = None
        # Auto-detect label column
        for col in df.columns:
            if col.lower() in ("label", "ground_truth", "actual", "true_label"):
                label_col = col
                break

    elif "file_data" in request.files and "file_pred" in request.files:
        mode1, data1 = load_uploaded_file(request.files["file_data"])
        mode2, data2 = load_uploaded_file(request.files["file_pred"])
        if mode1 != "dataframe" or mode2 != "dataframe":
            return jsonify({"error": "Both files must be CSV/JSON/XLSX"}), 400
        if len(data1) != len(data2):
            return jsonify({"error": f"Row count mismatch: data={len(data1)}, predictions={len(data2)}"}), 400
        df = pd.concat([data1, data2], axis=1)
        prediction_col = request.form.get("prediction_col", "").strip()
        if not prediction_col:
            prediction_col = detect_target_column(df) or df.columns[-1]
        label_col = request.form.get("label_col", "").strip() or None
    else:
        return jsonify({"error": "No file uploaded"}), 400

    metrics = compute_posttraining_metrics(df, prediction_col, label_col)
    if "error" in metrics:
        return jsonify({"error": metrics["error"]}), 400

    graphs = generate_graph_data(metrics, "post")

    groq_explanations = []
    for g in graphs:
        explanation = explain_graph_with_groq({"title": g["title"], "labels": g["labels"], "data": g["data"]})
        groq_explanations.append({"chart": g["title"], "explanation": explanation})

    post_alerts = []
    for attr, di in metrics.get("disparate_impact", {}).items():
        spd = metrics.get("statistical_parity_difference", {}).get(attr, 0)
        if di is not None and (di < 0.80 or (spd and spd > 0.10)):
            post_alerts.append({
                "type": "outcome_bias",
                "column": attr,
                "target": prediction_col,
                "statistical_parity_difference": spd,
                "disparate_impact": di,
                "severity": "HIGH" if (di < 0.60 or spd > 0.20) else "MEDIUM"
            })

    alert_explanations = []
    for alert in post_alerts:
        explanation = generate_alert_explanation_groq(alert, prediction_col)
        alert_explanations.append({
            "column": alert.get("column"),
            "type": alert.get("type"),
            "severity": alert.get("severity", "HIGH"),
            "explanation": explanation,
            "spd": alert.get("statistical_parity_difference"),
            "di": alert.get("disparate_impact"),
            "target": alert.get("target"),
        })

    report_result = generate_ai_report(metrics, "post")
    uploaded_file = request.files.get("file") or request.files.get("file_data")
    filename = uploaded_file.filename if uploaded_file else "unknown"

    result = {
        "metrics": metrics,
        "graphs": graphs,
        "groq_explanations": groq_explanations,
        "ai_explanations": groq_explanations,
        "alert_explanations": alert_explanations,
        "report": report_result.get("report", ""),
        "report_error": report_result.get("error"),
        "provider": report_result.get("provider", get_ai_provider()),
        "filename": secure_filename(filename) if filename else "unknown",
        "df_columns": list(df.columns),
        "auto_detected_prediction_col": prediction_col,
    }

    if MONGO_AVAILABLE:
        doc = normalize_for_mongo({
            "user_id": session["user_id"],
            "mode": "post-training",
            "filename": result["filename"],
            "metrics": metrics,
            "graphs": graphs,
            "groq_explanations": groq_explanations,
            "ai_explanations": groq_explanations,
            "alert_explanations": alert_explanations,
            "report": report_result.get("report", ""),
            "provider": report_result.get("provider", get_ai_provider()),
            "bias_score": metrics.get("bias_score", 0),
            "risk_level": metrics.get("risk_level", "UNKNOWN"),
            "created_at": datetime.now(timezone.utc).isoformat()
        })
        inserted = db["reports"].insert_one(doc)
        result["report_id"] = str(inserted.inserted_id)

    return jsonify(normalize_for_mongo(result))


# ─── What-If Routes ───────────────────────────────────────────────────────────
@app.route("/api/whatif/pre", methods=["POST"])
@login_required
def api_whatif_pre():
    data = request.get_json()
    if not data:
        return jsonify({"error": "No data sent"}), 400
    df_data = data.get("df_data")
    protected_attr = data.get("protected_attr", "")
    desired_balance = data.get("desired_balance", 50)
    if not df_data:
        return jsonify({"error": "No dataset data provided"}), 400
    try:
        df = pd.DataFrame(df_data)
    except Exception as e:
        return jsonify({"error": str(e)}), 400
    result = run_what_if_pretraining(df, protected_attr, desired_balance)
    if "error" not in result:
        result["groq_explanation"] = explain_optimizer_with_groq(result)
        result["ai_explanation"] = result["groq_explanation"]

    if MONGO_AVAILABLE and "error" not in result:
        try:
            doc = normalize_for_mongo({
                "user_id": session["user_id"],
                "mode": "whatif_pre",
                "filename": f"What-If Pre-Training Simulation - {protected_attr}",
                "file_name": f"What-If Pre-Training Simulation - {protected_attr}",
                "protected_attributes": [protected_attr] if protected_attr else [],
                "simulation_data": result,
                "desired_balance": desired_balance,
                "risk_level": "UNKNOWN",
                "bias_score": 0,
                "report": f"What-If Pre-Training Simulation\n\nProtected Attribute: {protected_attr}\nDesired Balance: {desired_balance}%\n\n{result.get('groq_explanation', '')}",
                "created_at": datetime.now(timezone.utc).isoformat()
            })
            inserted = db["reports"].insert_one(doc)
            result["report_id"] = str(inserted.inserted_id)
        except Exception as e:
            print(f"[What-If Pre] MongoDB error: {e}")
            pass

    return jsonify(normalize_for_mongo(result))

@app.route("/api/whatif/post", methods=["POST"])
@login_required
def api_whatif_post():
    data = request.get_json()
    if not data:
        return jsonify({"error": "No data sent"}), 400
    df_data = data.get("df_data")
    threshold = data.get("threshold", 0.5)
    fairness_weight = data.get("fairness_weight", 0.5)
    protected_attr = data.get("protected_attr", "")
    prediction_col = data.get("prediction_col", "").strip() or None
    if not df_data:
        return jsonify({"error": "No dataset data provided"}), 400
    try:
        df = pd.DataFrame(df_data)
    except Exception as e:
        return jsonify({"error": str(e)}), 400
    result = run_what_if_posttraining(df, threshold, fairness_weight, protected_attr, prediction_col)
    if "error" not in result:
        result["groq_explanation"] = explain_optimizer_with_groq(result)
        result["ai_explanation"] = result["groq_explanation"]

    if MONGO_AVAILABLE and "error" not in result:
        try:
            doc = normalize_for_mongo({
                "user_id": session["user_id"],
                "mode": "whatif_post",
                "filename": f"What-If Post-Training Simulation - {protected_attr}",
                "file_name": f"What-If Post-Training Simulation - {protected_attr}",
                "protected_attributes": [protected_attr] if protected_attr else [],
                "simulation_data": result,
                "threshold": threshold,
                "fairness_weight": fairness_weight,
                "risk_level": "UNKNOWN",
                "bias_score": 0,
                "report": f"What-If Post-Training Simulation\n\nProtected Attribute: {protected_attr}\nThreshold: {threshold}\nFairness Weight: {fairness_weight}",
                "created_at": datetime.now(timezone.utc).isoformat()
            })
            inserted = db["reports"].insert_one(doc)
            result["report_id"] = str(inserted.inserted_id)
        except Exception as e:
            print(f"[What-If Post] MongoDB error: {e}")
            pass

    return jsonify(normalize_for_mongo(result))

@app.route("/api/whatif/explain", methods=["POST"])
@login_required
def api_whatif_explain():
    data = request.get_json()
    if not data:
        return jsonify({"error": "No data sent"}), 400
    explanation = explain_optimizer_with_groq(data)

    if MONGO_AVAILABLE:
        try:
            doc = normalize_for_mongo({
                "user_id": session["user_id"],
                "simulation_type": "whatif_explain",
                "input": data,
                "explanation": explanation,
                "created_at": datetime.now(timezone.utc).isoformat()
            })
            db["simulations"].insert_one(doc)
        except Exception:
            pass

    return jsonify({"explanation": explanation})


# ─── Stress Testing Routes ────────────────────────────────────────────────────
@app.route("/api/stress/api", methods=["POST"])
@login_required
def api_stress_api():
    data = request.get_json()
    if not data:
        return jsonify({"error": "No data sent"}), 400
    result = run_blackbox_stress_test_api(
        data.get("api_url", ""), data.get("token", ""),
        data.get("sample_json", "{}"), data.get("protected_attr", "gender")
    )

    if MONGO_AVAILABLE:
        try:
            doc = normalize_for_mongo({
                "user_id": session["user_id"],
                "mode": "stress_test_api",
                "filename": f"Stress Test - API - {data.get('protected_attr', 'attribute')}",
                "file_name": f"Stress Test - API - {data.get('protected_attr', 'attribute')}",
                "protected_attributes": [data.get("protected_attr", "gender")],
                "api_url": data.get("api_url", ""),
                "simulation_data": result,
                "risk_level": "UNKNOWN",
                "bias_score": 0,
                "report": f"Stress Test - API\n\nAPI URL: {data.get('api_url', 'N/A')}\nProtected Attribute: {data.get('protected_attr', 'N/A')}\n\nTest Results: {result.get('summary', '')}",
                "created_at": datetime.now(timezone.utc).isoformat()
            })
            inserted = db["reports"].insert_one(doc)
            result["report_id"] = str(inserted.inserted_id)
        except Exception as e:
            print(f"[Stress API] MongoDB error: {e}")
            pass

    return jsonify(normalize_for_mongo(result))

@app.route("/api/stress/sandbox", methods=["POST"])
@login_required
def api_stress_sandbox():
    data = request.get_json()
    if not data:
        return jsonify({"error": "No data sent"}), 400
    result = run_blackbox_stress_test_sandbox(
        data.get("code", ""), data.get("sample_json", "{}"), data.get("protected_attr", "gender")
    )

    if MONGO_AVAILABLE:
        try:
            doc = normalize_for_mongo({
                "user_id": session["user_id"],
                "mode": "stress_test_sandbox",
                "filename": f"Stress Test - Sandbox - {data.get('protected_attr', 'attribute')}",
                "file_name": f"Stress Test - Sandbox - {data.get('protected_attr', 'attribute')}",
                "protected_attributes": [data.get("protected_attr", "gender")],
                "simulation_data": result,
                "risk_level": "UNKNOWN",
                "bias_score": 0,
                "report": f"Stress Test - Sandbox\n\nProtected Attribute: {data.get('protected_attr', 'N/A')}\n\nTest Results: {result.get('summary', '')}",
                "created_at": datetime.now(timezone.utc).isoformat()
            })
            inserted = db["reports"].insert_one(doc)
            result["report_id"] = str(inserted.inserted_id)
        except Exception as e:
            print(f"[Stress Sandbox] MongoDB error: {e}")
            pass

    return jsonify(normalize_for_mongo(result))

@app.route("/api/stress/generate", methods=["POST"])
@login_required
def api_stress_generate():
    data = request.get_json()
    if not data:
        return jsonify({"error": "No data sent"}), 400
    result = generate_sandbox_with_groq(data.get("protected_attr", "gender"), data.get("domain", "hiring"))

    if MONGO_AVAILABLE:
        try:
            doc = normalize_for_mongo({
                "user_id": session["user_id"],
                "mode": "stress_test_generate",
                "filename": f"Stress Test - Generated - {data.get('protected_attr', 'attribute')}",
                "file_name": f"Stress Test - Generated - {data.get('protected_attr', 'attribute')}",
                "protected_attributes": [data.get("protected_attr", "gender")],
                "domain": data.get("domain", "hiring"),
                "simulation_data": result,
                "risk_level": "UNKNOWN",
                "bias_score": 0,
                "report": f"Stress Test - Generated\n\nDomain: {data.get('domain', 'N/A')}\nProtected Attribute: {data.get('protected_attr', 'N/A')}\n\nGenerated Test Cases: {len(result.get('test_cases', []))}",
                "created_at": datetime.now(timezone.utc).isoformat()
            })
            inserted = db["reports"].insert_one(doc)
            result["report_id"] = str(inserted.inserted_id)
        except Exception as e:
            print(f"[Stress Generate] MongoDB error: {e}")
            pass

    return jsonify(result)


# ─── Appeal Engine Route ──────────────────────────────────────────────────────
@app.route("/api/appeal", methods=["POST"])
@login_required
def api_appeal():
    document_text = ""
    policy_text = ""
    domain = request.form.get("domain", "hiring")

    if "file_doc" in request.files and request.files["file_doc"].filename:
        mode, data = load_uploaded_file(request.files["file_doc"])
        if mode in ("text", "dataframe"):
            document_text = data if mode == "text" else data.to_string()
        else:
            return jsonify({"error": data}), 400
    else:
        document_text = request.form.get("doc_text", "").strip()

    policy_url = request.form.get("policy_url", "").strip()
    if policy_url:
        try:
            resp = http_requests.get(policy_url, timeout=10, headers={"User-Agent": "Mozilla/5.0"})
            policy_text = resp.text[:5000]
        except Exception as e:
            policy_text = f"Could not fetch URL: {e}"
    elif "file_policy" in request.files and request.files["file_policy"].filename:
        mode, data = load_uploaded_file(request.files["file_policy"])
        policy_text = data if mode in ("text", "dataframe") else ""
    else:
        policy_text = request.form.get("policy_text", "").strip()

    if not document_text:
        return jsonify({"error": "Document text is required"}), 400
    if not policy_text:
        return jsonify({"error": "Policy text or URL is required"}), 400

    result = run_appeal_engine(document_text, policy_text, domain)
    if "error" in result:
        return jsonify(result), 500

    if MONGO_AVAILABLE:
        doc = normalize_for_mongo({
            "user_id": session["user_id"],
            "mode": "appeal",
            "domain": domain,
            "document_preview": document_text[:500],
            "policy_preview": policy_text[:500],
            "report": result.get("report", ""),
            "fit_score": result.get("fit_score", 0),
            "provider": result.get("provider", ""),
            "bias_score": 100 - result.get("fit_score", 50),
            "risk_level": "MEDIUM",
            "filename": "appeal",
            "created_at": datetime.now(timezone.utc).isoformat()
        })
        inserted = db["reports"].insert_one(doc)
        result["report_id"] = str(inserted.inserted_id)

    return jsonify(result)


# ─── Reports Routes ───────────────────────────────────────────────────────────
@app.route("/api/reports")
@login_required
def api_reports():
    if not MONGO_AVAILABLE:
        return jsonify({"reports": [], "message": "MongoDB unavailable"})
    try:
        reports = list(db["reports"].find(
            {"user_id": session["user_id"]},
            {"report": 0, "metrics": 0, "graphs": 0}
        ).sort("created_at", -1).limit(50))
        for r in reports:
            r["_id"] = str(r["_id"])
        return jsonify({"reports": reports})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/reports/<report_id>")
@login_required
def api_report_detail(report_id):
    if not MONGO_AVAILABLE:
        return jsonify({"error": "MongoDB unavailable"}), 503
    try:
        report = db["reports"].find_one({"_id": ObjectId(report_id), "user_id": session["user_id"]})
        if not report:
            return jsonify({"error": "Report not found"}), 404
        report["_id"] = str(report["_id"])
        return jsonify(normalize_for_mongo(report))
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/reports/<report_id>/rename", methods=["PATCH"])
@login_required
def api_report_rename(report_id):
    if not MONGO_AVAILABLE:
        return jsonify({"error": "MongoDB unavailable"}), 503
    try:
        data = request.get_json()
        new_name = (data or {}).get("name", "").strip()
        if not new_name:
            return jsonify({"error": "Name cannot be empty"}), 400
        result = db["reports"].update_one(
            {"_id": ObjectId(report_id), "user_id": session["user_id"]},
            {"$set": {"filename": new_name}}
        )
        if result.matched_count == 0:
            return jsonify({"error": "Report not found"}), 404
        return jsonify({"ok": True, "filename": new_name})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/reports/<report_id>", methods=["DELETE"])
@login_required
def api_report_delete(report_id):
    if not MONGO_AVAILABLE:
        return jsonify({"error": "MongoDB unavailable"}), 503
    try:
        result = db["reports"].delete_one(
            {"_id": ObjectId(report_id), "user_id": session["user_id"]}
        )
        if result.deleted_count == 0:
            return jsonify({"error": "Report not found"}), 404
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


def _sanitize_json_value(val):
    if val is None:
        return ""
    if isinstance(val, float) and (np.isnan(val) or np.isinf(val)):
        return ""
    if isinstance(val, (np.integer,)):
        return int(val)
    if isinstance(val, (np.floating,)):
        if np.isnan(val) or np.isinf(val):
            return ""
        return float(val)
    if isinstance(val, (np.bool_,)):
        return bool(val)
    if isinstance(val, (pd.Timestamp,)):
        return str(val)
    return val


def _sanitize_profile(profile):
    return {str(k): _sanitize_json_value(v) for k, v in profile.items()}


def _counterfactual_base_key(profile, protected_attr):
    protected_lower = str(protected_attr).lower().strip()
    parts = []
    for k, v in sorted(profile.items()):
        if k == "_stress_type":
            continue
        if str(k).lower().strip() == protected_lower:
            continue
        parts.append(f"{k}={v}")
    stress = profile.get("_stress_type", "standard_qualification")
    return f"{stress}|" + "|".join(parts)


def _analyze_counterfactual_pairs(results, protected_attr):
    groups = {}
    for r in results:
        if "error" in r:
            continue
        profile = r.get("profile", {})
        key = _counterfactual_base_key(profile, protected_attr)
        groups.setdefault(key, []).append({
            "attr_val": _profile_attr_value(profile, protected_attr),
            "outcome": _normalize_prediction_outcome(r.get("result", 0)),
            "profile": profile,
            "stress_type": profile.get("_stress_type", "standard_qualification"),
        })

    pair_details = []
    flipped_profiles = []
    stable_profiles = []
    scenario_stats = {}

    for key, items in groups.items():
        if len(items) < 2:
            continue
        stress_type = items[0]["stress_type"]
        outcomes = [i["outcome"] for i in items]
        by_attr = {i["attr_val"]: i["outcome"] for i in items}
        flipped = len(set(outcomes)) > 1
        spread = max(outcomes) - min(outcomes)

        detail = {
            "stress_type": stress_type,
            "flipped": flipped,
            "outcomes_by_group": by_attr,
            "outcome_spread": spread,
        }
        pair_details.append(detail)

        if flipped:
            flipped_profiles.append(detail)
        else:
            stable_profiles.append(detail)

        scenario_stats.setdefault(stress_type, {"total": 0, "flipped": 0, "groups": {}})
        scenario_stats[stress_type]["total"] += 1
        if flipped:
            scenario_stats[stress_type]["flipped"] += 1
        for attr_val, outcome in by_attr.items():
            scenario_stats[stress_type]["groups"].setdefault(attr_val, {"approved": 0, "total": 0})
            scenario_stats[stress_type]["groups"][attr_val]["total"] += 1
            if outcome == 1:
                scenario_stats[stress_type]["groups"][attr_val]["approved"] += 1

    total_bases = len(pair_details)
    flip_count = len(flipped_profiles)
    flip_rate = round(flip_count / total_bases, 4) if total_bases > 0 else 0.0

    avg_spread = 0.0
    if flipped_profiles:
        avg_spread = sum(p["outcome_spread"] for p in flipped_profiles) / len(flipped_profiles)

    # Primary counterfactual score: based on how often changing only the protected attribute flips the decision
    if total_bases == 0:
        bias_score = 0.0
    else:
        bias_score = min(100.0, flip_rate * 100.0 + avg_spread * 35.0)

    max_gap_pct = round(avg_spread * 100, 1) if flipped_profiles else 0.0
    if flipped_profiles:
        all_group_rates = {}
        for p in flipped_profiles:
            for g, o in p["outcomes_by_group"].items():
                all_group_rates.setdefault(g, []).append(o)
        if len(all_group_rates) >= 2:
            rates = {g: sum(v) / len(v) for g, v in all_group_rates.items()}
            max_gap_pct = round((max(rates.values()) - min(rates.values())) * 100, 1)

    return {
        "total_base_profiles": total_bases,
        "flip_count": flip_count,
        "flip_rate": flip_rate,
        "flip_rate_pct": round(flip_rate * 100, 1),
        "stable_count": len(stable_profiles),
        "pair_details": pair_details,
        "flipped_profiles": flipped_profiles,
        "scenario_stats": scenario_stats,
        "counterfactual_bias_score": round(bias_score, 1),
        "max_outcome_spread_pct": max_gap_pct,
    }


def _auto_select_stress_attribute(df):
    df_binned = bin_age_column(_coerce_dataframe_types(df.copy()))
    attrs = detect_protected_attributes(df_binned)
    if not attrs:
        for col in df_binned.columns:
            col_lower = str(col).lower()
            if any(kw in col_lower for kw in ("gender", "race", "age", "sex", "ethnic")):
                attrs.append(col)
    if not attrs:
        target_col = detect_target_column(df_binned)
        for col in df_binned.columns:
            if col == target_col:
                continue
            try:
                unique_count = int(df_binned[col].dropna().astype(str).nunique())
            except Exception:
                unique_count = 0
            if 2 <= unique_count <= 20:
                return col, "fallback: categorical column with multiple groups"
    if not attrs:
        return None, "No protected attributes detected in dataset."

    target_col = detect_target_column(df_binned)
    best_attr = attrs[0]
    best_spd = -1.0
    selection_reason = "first detected protected attribute"

    if target_col and target_col in df_binned.columns:
        df_tmp = df_binned.copy()
        df_tmp["_stress_y_"] = convert_to_binary(df_tmp[target_col])
        for attr in attrs:
            if attr not in df_tmp.columns:
                continue
            try:
                grp_rates = df_tmp.groupby(df_tmp[attr].astype(str))["_stress_y_"].mean()
                if len(grp_rates) >= 2:
                    spd = float(grp_rates.max() - grp_rates.min())
                    if spd > best_spd:
                        best_spd = spd
                        best_attr = attr
                        selection_reason = f"highest outcome disparity (SPD={round(spd, 3)})"
            except Exception:
                pass

    return best_attr, selection_reason


def _build_stress_sections(protected_attr, pair_metrics, mode_label="dataset"):
    total = pair_metrics.get("total_base_profiles", 0)
    flip_count = pair_metrics.get("flip_count", 0)
    flip_pct = pair_metrics.get("flip_rate_pct", 0)
    score = pair_metrics.get("counterfactual_bias_score", 0)
    max_spread = pair_metrics.get("max_outcome_spread_pct", 0)

    if flip_count == 0:
        evidence = (
            f"Tested {total} matched base profiles. Changing only '{protected_attr}' never altered the "
            f"{mode_label} decision across any scenario."
        )
        sensitive = f"0/{total} profiles were counterfactually sensitive (0% flip rate)."
        verdict = f"NO counterfactual bias detected on '{protected_attr}' (score {score}/100, risk LOW)."
        recommendation = (
            f"Continue monitoring '{protected_attr}', but current {mode_label} decisions appear driven by "
            "qualification features (income, education, etc.) rather than this attribute alone."
        )
        finding = (
            f"'{protected_attr}' does not independently change decisions in this {mode_label}. "
            "Outcomes stay stable when only this attribute varies."
        )
    else:
        examples = []
        for item in pair_metrics.get("flipped_profiles", [])[:3]:
            parts = [f"{g}: {'approved' if o == 1 else 'rejected'}" for g, o in item["outcomes_by_group"].items()]
            examples.append(f"{item['stress_type']} → {', '.join(parts)}")
        evidence = (
            f"{flip_count}/{total} base profiles ({flip_pct}%) flipped decisions when only '{protected_attr}' changed. "
            f"Max outcome spread: {max_spread}%."
        )
        sensitive = f"Sensitive examples: {'; '.join(examples) if examples else 'see scenario breakdown'}."
        verdict = f"COUNTERFACTUAL BIAS detected on '{protected_attr}' (score {score}/100)."
        recommendation = (
            f"Review how '{protected_attr}' influences the {mode_label}. Consider fairness constraints, "
            "threshold adjustment, or retraining with bias mitigation."
        )
        finding = (
            f"Changing '{protected_attr}' alone changes decisions in {flip_pct}% of tested profiles."
        )

    return {
        "KEY FINDING": finding,
        "COUNTERFACTUAL EVIDENCE": evidence,
        "SENSITIVE PROFILES": sensitive,
        "VERDICT": verdict,
        "RECOMMENDATION": recommendation,
    }


def _parse_stress_sections(text):
    labels = ["KEY FINDING", "COUNTERFACTUAL EVIDENCE", "SENSITIVE PROFILES", "VERDICT", "RECOMMENDATION"]
    sections = {}
    if not text:
        return sections
    current = None
    buffer = []
    for line in text.split("\n"):
        stripped = line.strip()
        if not stripped:
            continue
        matched = None
        for label in labels:
            if stripped.upper().startswith(label + ":") or stripped.upper() == label + ":":
                if current and buffer:
                    sections[current] = " ".join(buffer).strip()
                current = label
                rest = stripped.split(":", 1)[1].strip() if ":" in stripped else ""
                buffer = [rest] if rest else []
                matched = label
                break
        if matched is None and current:
            buffer.append(stripped)
    if current and buffer:
        sections[current] = " ".join(buffer).strip()
    return sections


def _build_rule_based_stress_explanation(protected_attr, pair_metrics, mode_label="dataset"):
    sections = _build_stress_sections(protected_attr, pair_metrics, mode_label)
    return "\n\n".join([f"{k}:\n{v}" for k, v in sections.items()])


def analyze_stress_results(results, protected_attr, mode_label="dataset"):
    pair_metrics = _analyze_counterfactual_pairs(results, protected_attr)
    scenario_stats = pair_metrics.get("scenario_stats", {})

    # Build summary focused on counterfactual sensitivity per scenario
    summary = {}
    for stress_type, stats in scenario_stats.items():
        summary[stress_type] = {
            "_meta": {
                "base_profiles": stats.get("total", 0),
                "sensitive_profiles": stats.get("flipped", 0),
                "flip_rate_pct": round((stats.get("flipped", 0) / stats["total"]) * 100, 1) if stats.get("total", 0) > 0 else 0.0,
                "counterfactual_sensitive": stats.get("flipped", 0) > 0,
            }
        }
        for attr_val, counts in stats.get("groups", {}).items():
            summary[stress_type][attr_val] = {
                "approved": counts.get("approved", 0),
                "total": counts.get("total", 0),
                "rate_pct": round((counts["approved"] / counts["total"]) * 100, 1) if counts.get("total", 0) > 0 else 0.0,
            }

    overall_bias_score = pair_metrics.get("counterfactual_bias_score", 0)
    max_gap_pct = pair_metrics.get("max_outcome_spread_pct", 0)

    metrics_payload = {
        "counterfactual_flip_count": pair_metrics.get("flip_count", 0),
        "counterfactual_base_profiles": pair_metrics.get("total_base_profiles", 0),
        "counterfactual_flip_rate_pct": pair_metrics.get("flip_rate_pct", 0),
        "counterfactual_stable_profiles": pair_metrics.get("stable_count", 0),
        "overall_bias_score": overall_bias_score,
        "overall_risk_level": score_to_risk(overall_bias_score),
        "max_approval_gap_pct": max_gap_pct,
        "pair_analysis": pair_metrics.get("pair_details", []),
    }

    prompt = f"""You are a senior AI fairness auditor. Analyze these TRUE counterfactual stress test results.
PROTECTED ATTRIBUTE TESTED: {protected_attr}
MODE: {mode_label}

IMPORTANT: For each base profile, ONLY '{protected_attr}' was changed while all other features stayed identical.

COUNTERFACTUAL METRICS:
- Base profiles tested: {pair_metrics.get('total_base_profiles', 0)}
- Sensitive profiles (decision changed): {pair_metrics.get('flip_count', 0)}
- Flip rate: {pair_metrics.get('flip_rate_pct', 0)}%
- Counterfactual bias score: {overall_bias_score}/100
- Max outcome spread in sensitive profiles: {max_gap_pct}%

SCENARIO BREAKDOWN:
{json.dumps({k: v.get('_meta', {}) for k, v in summary.items()}, indent=2)}

FLIPPED PROFILE EXAMPLES:
{json.dumps(pair_metrics.get('flipped_profiles', [])[:5], indent=2)}

Return EXACTLY these 5 sections (1-2 short sentences each, cite exact numbers):

KEY FINDING:
COUNTERFACTUAL EVIDENCE:
SENSITIVE PROFILES:
VERDICT:
RECOMMENDATION:

STRICT RULES:
- If flip_count is 0, say clearly that '{protected_attr}' alone does NOT change decisions.
- If flip_count > 0, cite exact flip counts and examples.
- Do NOT write long paragraphs. Keep each section concise.
- Do not confuse aggregate group rates with counterfactual sensitivity.
"""
    ai_text, groq_error = call_groq_for_stress(prompt, max_tokens=700, temperature=0.2)
    ai_text = ai_text or ""
    sections = _parse_stress_sections(ai_text)
    if len(sections) < 3:
        sections = _build_stress_sections(protected_attr, pair_metrics, mode_label)
    explanation = "\n\n".join([f"{k}:\n{v}" for k, v in sections.items()])

    return {
        "explanation": explanation,
        "sections": sections,
        "summary": summary,
        "metrics": metrics_payload,
        "counterfactual_bias_score": overall_bias_score,
        "risk_level": score_to_risk(overall_bias_score),
        "pair_metrics": pair_metrics,
        "ai_provider": "groq" if ai_text and not groq_error else "groq_timeout_fallback",
        "ai_warning": groq_error or "",
    }


def _select_surrogate_target_column(df, protected_attr):
    target_col = detect_target_column(df)
    if target_col and target_col in df.columns:
        return target_col
    resolved_attr = _resolve_column_name(df, protected_attr) or protected_attr
    for col in df.columns:
        if col != resolved_attr:
            return col
    return None


def _profile_column_value(profile, col):
    if col in profile:
        return profile.get(col)
    lookup = {str(k).lower().strip(): v for k, v in profile.items()}
    return lookup.get(str(col).lower().strip(), "")


def _profile_distance(profile, row, feature_columns, template_df):
    distance = 0.0
    compared = 0
    for col in feature_columns:
        left = _profile_column_value(profile, col)
        right = _profile_column_value(row, col)
        if pd.isna(left) or pd.isna(right) or left == "" or right == "":
            continue
        compared += 1
        if col in template_df.columns and pd.api.types.is_numeric_dtype(template_df[col]):
            try:
                series = pd.to_numeric(template_df[col], errors="coerce")
                scale = float(series.max() - series.min()) or 1.0
                distance += abs(float(left) - float(right)) / scale
            except Exception:
                distance += 1.0
        else:
            distance += 0.0 if str(left).lower().strip() == str(right).lower().strip() else 1.0
    if compared == 0:
        return float("inf")
    return distance


def run_lightweight_baseline_predictions(df, profiles, protected_attr):
    """Pure-Pandas fallback for Render environments where sklearn is unavailable."""
    df_model = bin_age_column(_coerce_dataframe_types(df.copy()))
    target_col = _select_surrogate_target_column(df_model, protected_attr)
    if not target_col or target_col not in df_model.columns:
        return None, "Could not find a target or prediction column for stress testing."

    y = convert_to_binary(df_model[target_col])
    feature_columns = [c for c in df_model.columns if c != target_col]
    if not feature_columns:
        default_pred = int(round(float(y.mean()))) if len(y) else 0
        return [{"profile": row, "result": default_pred} for row in profiles], None

    rows = df_model[feature_columns].to_dict(orient="records")
    labels = y.tolist()
    default_pred = int(round(float(y.mean()))) if len(y) else 0
    results = []

    for row in profiles:
        safe_row = _sanitize_profile(row)
        try:
            best_idx = None
            best_dist = float("inf")
            for idx, train_row in enumerate(rows):
                dist = _profile_distance(safe_row, train_row, feature_columns, df_model)
                if dist < best_dist:
                    best_dist = dist
                    best_idx = idx
            pred = int(labels[best_idx]) if best_idx is not None and best_dist != float("inf") else default_pred
            results.append({"profile": safe_row, "result": pred})
        except Exception as e:
            results.append({"profile": safe_row, "result": default_pred, "warning": str(e)})

    return results, None


def run_baseline_predictions(df, profiles, protected_attr):
    resolved_attr = _resolve_column_name(df, protected_attr) or protected_attr
    df_model = bin_age_column(_coerce_dataframe_types(df.copy()))
    return run_lightweight_baseline_predictions(df_model, profiles, resolved_attr)


def run_dataset_stress_test(df_data, protected_attr=None, mode="pre", code=""):
    try:
        df = _coerce_dataframe_types(pd.DataFrame(df_data))
    except Exception as e:
        return {"error": f"Invalid dataset: {str(e)}"}

    if len(df) > 5000:
        df = df.sample(n=5000, random_state=42)

    if len(df) < 5:
        return {"error": "Dataset too small for stress testing (need at least 5 rows)."}

    selection_reason = ""
    if not protected_attr or str(protected_attr).strip().lower() in ("", "auto"):
        auto_attr, selection_reason = _auto_select_stress_attribute(df)
        if not auto_attr:
            return {"error": selection_reason or "Could not auto-detect a protected attribute."}
        protected_attr = auto_attr

    resolved_attr = _resolve_column_name(df, protected_attr)
    if not resolved_attr:
        available = ", ".join(str(c) for c in df.columns)
        return {"error": f"No protected attribute found. Available columns: {available}"}

    grid = [_sanitize_profile(p) for p in generate_counterfactual_grid(df, resolved_attr)]
    if not grid:
        unique_vals = df[resolved_attr].dropna().astype(str).unique().tolist() if resolved_attr in df.columns else []
        if len(unique_vals) < 2:
            return {
                "error": (
                    f"'{resolved_attr}' must contain at least 2 distinct groups for counterfactual testing. "
                    f"Found: {unique_vals or 'no values'}."
                )
            }
        return {"error": f"Could not generate counterfactual profiles for '{resolved_attr}'."}

    # Both pre and post use AI-trained surrogate model from the dataset (no manual code needed)
    pred_results, baseline_err = run_baseline_predictions(df, grid, resolved_attr)
    if baseline_err:
        return {"error": baseline_err}
    pred_results = pred_results or []

    if not pred_results:
        return {"error": "No prediction results returned from stress test."}

    success_count = sum(1 for r in pred_results if "error" not in r)
    if success_count == 0:
        sample_err = next((r.get("error") for r in pred_results if r.get("error")), "Unknown error")
        return {"error": f"All counterfactual profiles failed during prediction. Sample error: {sample_err}"}

    mode_label = "pre-training dataset model" if mode == "pre" else "post-training prediction model"
    analysis = analyze_stress_results(pred_results, resolved_attr, mode_label=mode_label)
    analysis["results"] = pred_results
    analysis["protected_attr"] = resolved_attr
    analysis["auto_selected_attribute"] = resolved_attr
    analysis["selection_reason"] = selection_reason or "user-specified attribute"
    analysis["profiles_tested"] = len(pred_results)
    analysis["successful_predictions"] = success_count
    return analysis


def run_sandbox_on_profiles(code, profiles):
    if not code or not code.strip():
        return {"error": "No code provided"}
    if not profiles:
        return {"error": "No profiles provided"}

    ALLOWED_BUILTINS = {
        "print": print, "range": range, "len": len, "int": int, "float": float,
        "str": str, "bool": bool, "list": list, "dict": dict, "abs": abs,
        "round": round, "max": max, "min": min, "sum": sum, "enumerate": enumerate,
        "zip": zip, "isinstance": isinstance, "hasattr": hasattr, "getattr": getattr,
        "type": type,
    }

    try:
        ns = {"__builtins__": ALLOWED_BUILTINS}
        exec(code, ns)
        predict = ns.get("predict")
        if predict is None:
            return {"error": "No predict() function found. Define: def predict(row): ..."}

        results = []
        for row in profiles:
            safe_row = _sanitize_profile(row)
            try:
                result = predict(safe_row)
                results.append({"profile": safe_row, "result": _normalize_prediction_outcome(result)})
            except Exception as e:
                results.append({"profile": safe_row, "error": str(e)})
        return {"results": results}
    except Exception as e:
        return {"error": f"Sandbox execution failed: {str(e)}"}


def train_baseline_model(df, protected_attr):
    from sklearn.linear_model import LogisticRegression
    from sklearn.preprocessing import OneHotEncoder
    from sklearn.compose import ColumnTransformer
    from sklearn.pipeline import Pipeline
    from sklearn.impute import SimpleImputer
    
    df = _coerce_dataframe_types(df.copy())
    resolved_attr = _resolve_column_name(df, protected_attr) or protected_attr
    target_col = detect_target_column(df)
    if not target_col:
        for c in df.columns:
            if _resolve_column_name(df, c) != resolved_attr and c != resolved_attr:
                target_col = c
                break
                
    if not target_col or target_col not in df.columns:
        return None
        
    X = df.drop(columns=[target_col])
    y = convert_to_binary(df[target_col])
    if y.nunique() < 2:
        return None
    
    numeric_features = X.select_dtypes(include=["number", "bool"]).columns.tolist()
    categorical_features = [c for c in X.columns if c not in numeric_features]
    
    transformers = []
    if numeric_features:
        transformers.append(
            ("num", Pipeline([("imputer", SimpleImputer(strategy="median"))]), numeric_features)
        )
    if categorical_features:
        transformers.append(
            ("cat", Pipeline([
                ("imputer", SimpleImputer(strategy="most_frequent")),
                ("onehot", OneHotEncoder(handle_unknown="ignore")),
            ]), categorical_features)
        )
    if not transformers:
        return None

    preprocessor = ColumnTransformer(transformers=transformers)
    
    clf = Pipeline(steps=[("preprocessor", preprocessor),
                          ("classifier", LogisticRegression(random_state=42, max_iter=500))])
    
    try:
        clf.fit(X, y)
        return clf
    except Exception:
        return None


def generate_counterfactual_grid(df, protected_attr):
    resolved_attr = _resolve_column_name(df, protected_attr)
    if not resolved_attr:
        return []

    df = _coerce_dataframe_types(bin_age_column(df))
    unique_vals = [str(v) for v in df[resolved_attr].dropna().unique().tolist()]
    if not unique_vals or len(unique_vals) < 2:
        return []

    target_col = detect_target_column(df)
    features_df = df.drop(columns=[target_col]) if target_col and target_col in df.columns else df.copy()
    features_df = features_df.drop(columns=[resolved_attr]) if resolved_attr in features_df.columns else features_df

    grid_profiles = []

    # ── STRATEGY 1: True Counterfactuals from real dataset rows ────────────────
    sample_size = min(5, len(df))
    try:
        sampled_rows = df.sample(n=sample_size, random_state=42)
    except Exception:
        sampled_rows = df.head(sample_size)

    for idx, row in sampled_rows.iterrows():
        base = {}
        for col in features_df.columns:
            val = row[col]
            if pd.isna(val):
                base[col] = ""
            elif isinstance(val, (int, float, complex)) and not isinstance(val, bool):
                base[col] = float(val)
            else:
                base[col] = str(val)

        for attr_val in unique_vals:
            profile = dict(base)
            profile[resolved_attr] = attr_val
            profile["_stress_type"] = f"real_row_{idx}"
            grid_profiles.append(profile)

    # ── STRATEGY 2: Synthetic tiers at low / median / high qualification ───────
    tier_specs = [
        ("low_qualification", 0.25),
        ("standard_qualification", 0.50),
        ("high_qualification", 0.75),
    ]

    for tier_name, q in tier_specs:
        baseline = {}
        for col in features_df.columns:
            col_data = features_df[col].dropna()
            if len(col_data) == 0:
                baseline[col] = ""
                continue
            if pd.api.types.is_numeric_dtype(col_data):
                baseline[col] = float(col_data.quantile(q))
            else:
                as_num = pd.to_numeric(col_data.astype(str).str.replace(",", "", regex=False), errors="coerce")
                if int(as_num.notna().sum()) >= max(1, int(len(col_data) * 0.8)):
                    baseline[col] = float(as_num.quantile(q))
                else:
                    mode_val = col_data.mode()
                    baseline[col] = str(mode_val.iloc[0]) if not mode_val.empty else ""

        for attr_val in unique_vals:
            profile = dict(baseline)
            profile[resolved_attr] = attr_val
            profile["_stress_type"] = tier_name
            grid_profiles.append(profile)

    return grid_profiles


@app.route("/api/stress/generate_adversarial_grid", methods=["POST"])
@login_required
def api_stress_generate_adversarial_grid():
    data = request.get_json()
    if not data:
        return jsonify({"error": "No data sent"}), 400
    df_data = data.get("df_data")
    protected_attr = data.get("protected_attr", "gender")
    if not df_data:
        return jsonify({"error": "No dataset data provided"}), 400
    try:
        df = pd.DataFrame(df_data)
    except Exception as e:
        return jsonify({"error": str(e)}), 400

    resolved_attr = _resolve_column_name(df, protected_attr)
    if not resolved_attr:
        available = ", ".join(str(c) for c in df.columns)
        return jsonify({
            "error": f"Protected attribute '{protected_attr}' not found. Available columns: {available}"
        }), 400
        
    grid = [_sanitize_profile(p) for p in generate_counterfactual_grid(df, resolved_attr)]
    return jsonify({"grid": grid, "protected_attr": resolved_attr})


@app.route("/api/stress/run_sandbox_profiles", methods=["POST"])
@login_required
def api_stress_run_sandbox_profiles():
    data = request.get_json()
    if not data:
        return jsonify({"error": "No data sent"}), 400
    code = data.get("code", "")
    profiles = data.get("profiles", [])
    if not code:
        return jsonify({"error": "No code provided"}), 400
    if not profiles:
        return jsonify({"error": "No profiles provided"}), 400
        
    res = run_sandbox_on_profiles(code, profiles)
    return jsonify(res)


@app.route("/api/stress/run_baseline_profiles", methods=["POST"])
@login_required
def api_stress_run_baseline_profiles():
    data = request.get_json()
    if not data:
        return jsonify({"error": "No data sent"}), 400
    df_data = data.get("df_data")
    profiles = data.get("profiles", [])
    protected_attr = data.get("protected_attr", "gender")
    if not df_data:
        return jsonify({"error": "No dataset data provided"}), 400
    if not profiles:
        return jsonify({"error": "No profiles provided"}), 400
        
    try:
        df = _coerce_dataframe_types(pd.DataFrame(df_data))
    except Exception as e:
        return jsonify({"error": str(e)}), 400

    resolved_attr = _resolve_column_name(df, protected_attr) or protected_attr
    results, err = run_baseline_predictions(df, [_sanitize_profile(p) for p in profiles], resolved_attr)
    if err:
        return jsonify({"error": err}), 400
    return jsonify({"results": results, "protected_attr": resolved_attr})


def _load_stress_dataframe_from_request():
    if "file_data" in request.files and "file_pred" in request.files:
        mode1, data1 = load_uploaded_file(request.files["file_data"])
        mode2, data2 = load_uploaded_file(request.files["file_pred"])
        if mode1 != "dataframe" or mode2 != "dataframe":
            return None, "Both stress-test files must be CSV/JSON/XLSX."
        if len(data1) != len(data2):
            return None, f"Row count mismatch: data={len(data1)}, predictions={len(data2)}."
        return pd.concat([data1.reset_index(drop=True), data2.reset_index(drop=True)], axis=1), None

    if "file" in request.files and request.files["file"].filename:
        mode, data = load_uploaded_file(request.files["file"])
        if mode == "error":
            return None, data
        if mode != "dataframe":
            return None, "Stress testing requires a structured CSV, JSON, or XLSX file."
        return data, None

    return None, "No stress-test file uploaded."


def _light_clean_value(value):
    if value is None:
        return ""
    text = str(value).strip()
    return "" if text.lower() in ("nan", "none", "null") else text


def _light_parse_uploaded_rows(file):
    filename = secure_filename(file.filename or "").lower()
    ext = filename.rsplit(".", 1)[-1] if "." in filename else "csv"

    if ext == "json":
        try:
            raw = json.loads(file.read().decode("utf-8", errors="replace"))
            rows = raw.get("data", raw.get("records", [])) if isinstance(raw, dict) else raw
            if not isinstance(rows, list):
                return None, "JSON stress file must be a list of records."
            return [{str(k): _light_clean_value(v) for k, v in r.items()} for r in rows if isinstance(r, dict)], None
        except Exception as e:
            return None, f"Could not parse JSON stress file: {str(e)}"

    if ext in ("xlsx", "xls"):
        try:
            from openpyxl import load_workbook
            wb = load_workbook(file, read_only=True, data_only=True)
            ws = wb.active
            values = list(ws.iter_rows(values_only=True))
            if len(values) < 2:
                return None, "Spreadsheet must include a header row and at least one data row."
            headers = [str(h).strip() if h is not None else f"column_{i+1}" for i, h in enumerate(values[0])]
            rows = []
            for vals in values[1:]:
                rows.append({headers[i]: _light_clean_value(vals[i] if i < len(vals) else "") for i in range(len(headers))})
            return rows, None
        except Exception as e:
            return None, f"Could not parse spreadsheet stress file: {str(e)}"

    try:
        text = file.read().decode("utf-8-sig", errors="replace")
        sample = text[:4096]
        try:
            dialect = csv.Sniffer().sniff(sample)
        except Exception:
            dialect = csv.excel
        reader = csv.DictReader(StringIO(text), dialect=dialect)
        rows = [{str(k): _light_clean_value(v) for k, v in row.items() if k is not None} for row in reader]
        return rows, None
    except Exception as e:
        return None, f"Could not parse CSV stress file: {str(e)}"


def _light_load_stress_rows_from_request():
    if "file_data" in request.files and "file_pred" in request.files:
        rows1, err1 = _light_parse_uploaded_rows(request.files["file_data"])
        rows2, err2 = _light_parse_uploaded_rows(request.files["file_pred"])
        if err1 or err2:
            return None, err1 or err2
        if len(rows1) != len(rows2):
            return None, f"Row count mismatch: data={len(rows1)}, predictions={len(rows2)}."
        return [dict(a, **b) for a, b in zip(rows1, rows2)], None

    if "file" not in request.files or not request.files["file"].filename:
        return None, "No stress-test file uploaded."
    return _light_parse_uploaded_rows(request.files["file"])


def _light_as_float(value):
    try:
        text = str(value).replace(",", "").strip()
        return float(text) if text != "" else None
    except Exception:
        return None


def _light_detect_target_col(rows):
    if not rows:
        return None
    keys = list(rows[0].keys())
    for key in keys:
        if str(key).lower().strip() in TARGET_KEYWORDS:
            return key
    return keys[-1] if keys else None


def _light_detect_attr(rows, target_col):
    if not rows:
        return None, "No rows available."
    keys = [k for k in rows[0].keys() if k != target_col]
    candidates = []
    for key in keys:
        values = [str(r.get(key, "")).strip() for r in rows if str(r.get(key, "")).strip()]
        unique = sorted(set(values))
        lower = str(key).lower()
        if any(kw in lower for kw in PROTECTED_KEYWORDS) and 2 <= len(unique) <= 30:
            candidates.append(key)
    if not candidates:
        for key in keys:
            values = [str(r.get(key, "")).strip() for r in rows if str(r.get(key, "")).strip()]
            unique = sorted(set(values))
            numeric_count = sum(1 for v in values if _light_as_float(v) is not None)
            if 2 <= len(unique) <= 20 and numeric_count < max(1, int(len(values) * 0.8)):
                return key, "fallback: categorical column with multiple groups"
        return None, "No protected or categorical attribute detected for stress testing."

    best_attr = candidates[0]
    best_gap = -1.0
    for attr in candidates:
        groups = {}
        for row in rows:
            group = str(row.get(attr, "")).strip()
            if not group:
                continue
            groups.setdefault(group, []).append(_normalize_prediction_outcome(row.get(target_col, 0)))
        rates = [sum(vals) / len(vals) for vals in groups.values() if vals]
        if len(rates) >= 2:
            gap = max(rates) - min(rates)
            if gap > best_gap:
                best_gap = gap
                best_attr = attr
    reason = f"highest outcome disparity (SPD={round(best_gap, 3)})" if best_gap >= 0 else "first detected protected attribute"
    return best_attr, reason


def _light_distance(a, b, feature_cols):
    distance = 0.0
    compared = 0
    for col in feature_cols:
        av = a.get(col, "")
        bv = b.get(col, "")
        if av == "" or bv == "":
            continue
        compared += 1
        af = _light_as_float(av)
        bf = _light_as_float(bv)
        if af is not None and bf is not None:
            scale = max(abs(af), abs(bf), 1.0)
            distance += abs(af - bf) / scale
        else:
            distance += 0.0 if str(av).lower().strip() == str(bv).lower().strip() else 1.0
    return distance if compared else float("inf")


def _light_predict_counterfactuals(rows, profiles, target_col):
    feature_cols = [c for c in rows[0].keys() if c != target_col]
    labels = [_normalize_prediction_outcome(row.get(target_col, 0)) for row in rows]
    default_pred = int(round(sum(labels) / len(labels))) if labels else 0
    results = []
    for profile in profiles:
        best_idx = None
        best_dist = float("inf")
        for idx, row in enumerate(rows):
            dist = _light_distance(profile, row, feature_cols)
            if dist < best_dist:
                best_dist = dist
                best_idx = idx
        pred = labels[best_idx] if best_idx is not None and best_dist != float("inf") else default_pred
        results.append({"profile": profile, "result": pred})
    return results


def run_light_stress_test_from_rows(rows, protected_attr="auto", mode="pre"):
    rows = [r for r in rows if isinstance(r, dict)]
    if len(rows) > 1000:
        step = len(rows) / 1000
        rows = [rows[int(i * step)] for i in range(1000)]
    if len(rows) < 5:
        return {"error": "Dataset too small for stress testing (need at least 5 rows)."}

    target_col = _light_detect_target_col(rows)
    if not target_col:
        return {"error": "Could not find a target or prediction column for stress testing."}

    if not protected_attr or str(protected_attr).strip().lower() in ("", "auto"):
        resolved_attr, selection_reason = _light_detect_attr(rows, target_col)
    else:
        lookup = {str(k).lower().strip(): k for k in rows[0].keys()}
        resolved_attr = lookup.get(str(protected_attr).lower().strip())
        selection_reason = "user-specified attribute"
    if not resolved_attr:
        return {"error": selection_reason or "Could not auto-detect a protected attribute."}

    attr_values = sorted({str(r.get(resolved_attr, "")).strip() for r in rows if str(r.get(resolved_attr, "")).strip()})
    if len(attr_values) < 2:
        return {"error": f"'{resolved_attr}' must contain at least 2 distinct groups for counterfactual testing."}

    sample_rows = rows[:5]
    profiles = []
    for idx, row in enumerate(sample_rows):
        base = {k: _light_clean_value(v) for k, v in row.items() if k != target_col}
        for attr_val in attr_values:
            profile = dict(base)
            profile[resolved_attr] = attr_val
            profile["_stress_type"] = f"real_row_{idx}"
            profiles.append(profile)

    pred_results = _light_predict_counterfactuals(rows, profiles, target_col)
    mode_label = "pre-training dataset model" if mode == "pre" else "post-training prediction model"
    analysis = analyze_stress_results(pred_results, resolved_attr, mode_label=mode_label)
    if analysis.get("error"):
        return analysis
    analysis["results"] = pred_results
    analysis["protected_attr"] = resolved_attr
    analysis["auto_selected_attribute"] = resolved_attr
    analysis["selection_reason"] = selection_reason
    analysis["profiles_tested"] = len(pred_results)
    analysis["successful_predictions"] = len(pred_results)
    return analysis


@app.route("/api/stress/run_dataset_test_light", methods=["POST"])
@login_required
def api_stress_run_dataset_test_light():
    try:
        rows, err = _light_load_stress_rows_from_request()
        if err:
            return jsonify({"error": err}), 400
        protected_attr = request.form.get("protected_attr", "auto")
        mode = request.form.get("mode", "pre")
        result = run_light_stress_test_from_rows(rows, protected_attr=protected_attr, mode=mode)
        if result.get("error"):
            return jsonify(result), 400
        return jsonify(normalize_for_mongo(result))
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": f"Light stress test failed: {type(e).__name__}: {str(e)}"}), 500


@app.route("/api/stress/run_dataset_test", methods=["POST"])
@login_required
def api_stress_run_dataset_test():
    try:
        if request.files:
            df_data, load_error = _load_stress_dataframe_from_request()
            if load_error:
                return jsonify({"error": load_error}), 400
            protected_attr = request.form.get("protected_attr", "auto")
            mode = request.form.get("mode", "pre")
            code = request.form.get("code", "")
        else:
            data = request.get_json(silent=True)
            if not data:
                return jsonify({"error": "No data sent"}), 400
            df_data = data.get("df_data")
            protected_attr = data.get("protected_attr", "gender")
            mode = data.get("mode", "pre")
            code = data.get("code", "")

        if df_data is None or (not isinstance(df_data, pd.DataFrame) and not df_data):
            return jsonify({"error": "No dataset data provided"}), 400

        result = run_dataset_stress_test(df_data, protected_attr, mode=mode, code=code)
        if result.get("error"):
            return jsonify(result), 400
        try:
            return jsonify(normalize_for_mongo(result))
        except Exception as e:
            traceback.print_exc()
            compact = {
                "explanation": result.get("explanation", ""),
                "sections": result.get("sections", {}),
                "summary": result.get("summary", {}),
                "metrics": result.get("metrics", {}),
                "counterfactual_bias_score": result.get("counterfactual_bias_score", 0),
                "risk_level": result.get("risk_level", "UNKNOWN"),
                "protected_attr": result.get("protected_attr", protected_attr),
                "auto_selected_attribute": result.get("auto_selected_attribute", protected_attr),
                "selection_reason": result.get("selection_reason", ""),
                "profiles_tested": result.get("profiles_tested", 0),
                "successful_predictions": result.get("successful_predictions", 0),
                "results": result.get("results", [])[:50],
                "warning": f"Returned compact stress result after serialization fallback: {type(e).__name__}",
            }
            return jsonify(normalize_for_mongo(compact))
    except Exception as e:
        traceback.print_exc()
        return jsonify({
            "error": f"Stress test failed on server: {type(e).__name__}: {str(e)}"
        }), 500


@app.route("/api/stress/analyze_grid_outcomes", methods=["POST"])
@login_required
def api_stress_analyze_grid_outcomes():
    data = request.get_json()
    if not data:
        return jsonify({"error": "No data sent"}), 400
    results = data.get("results", [])
    protected_attr = data.get("protected_attr", "gender")
    if not results:
        return jsonify({"error": "No stress test results provided"}), 400

    analysis = analyze_stress_results(results, protected_attr)
    return jsonify(normalize_for_mongo(analysis))


@app.errorhandler(RequestEntityTooLarge)
def handle_request_too_large(e):
    if request.path.startswith("/api/"):
        return jsonify({
            "error": "Uploaded data is too large for the server. Use a smaller file or fewer rows for stress testing."
        }), 413
    return e


@app.errorhandler(HTTPException)
def handle_api_http_exception(e):
    if request.path.startswith("/api/"):
        return jsonify({"error": e.description or e.name}), e.code
    return e


@app.errorhandler(Exception)
def handle_api_exception(e):
    if request.path.startswith("/api/"):
        traceback.print_exc()
        return jsonify({
            "error": f"Server error: {type(e).__name__}: {str(e)}"
        }), 500
    raise e


if __name__ == "__main__":
    os.makedirs("uploads", exist_ok=True)
    app.run(debug=True, port=5000)
